"""
Import Excel d'accords de mobilite.

Colonnes du template :
  A - Etablissement externe (nom universite partenaire)
  B - Nom de l'accord
  C - Departements concernes (codes separes par point-virgule, ex: SN;3EA)
  D - Niveaux concernes (codes, ex: 3A;2A)
  E - Cadre d'accord (nom du cadre)
  F - Nombre de places (entier ou 'illimite')
  G - Etablissements internes (noms courts, ex: INP-ENSEEIHT;INP-ENSAT)
  H - Remarques
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from itertools import combinations
from typing import Any

import openpyxl
import openpyxl.worksheet.datavalidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

N7_IDENTIFIER = "ENSEEIHT"

TEMPLATE_COLUMNS = [
    "Etablissement externe",
    "Nom de l'accord",
    "Departements concernes (codes, ex: SN;3EA)",
    "Niveaux concernes (codes, ex: 3A;2A)",
    "Cadre d'accord",
    "Nombre de places (entier ou 'illimite')",
    "Etablissements internes (nom court, ex: INP-ENSEEIHT;INP-ENSAT)",
    "Remarques",
]

COLUMN_MAP = {
    "etablissement_externe": 0,
    "nom": 1,
    "departements": 2,
    "niveaux": 3,
    "cadre": 4,
    "places": 5,
    "etablissements_internes": 6,
    "remarques": 7,
}

UNLIMITED_PLACES_VALUES = {
    "illimite",
    "illimité",
    "pas de lim",
    "pas de limite",
    "unlimited",
    "∞",
    "infini",
    "n/a",
    "-",
    "",
}


@dataclass
class ExcelRow:
    row_number: int
    raw: dict[str, Any]
    university_name: str = ""
    agreement_name: str = ""
    departments_raw: str = ""
    levels_raw: str = ""
    framework_raw: str = ""
    places: int | None = None
    institutions_raw: str = ""
    remarks: str = ""
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    @property
    def department_codes(self) -> list[str]:
        if not self.departments_raw:
            return []
        return [
            code.strip().upper()
            for code in self.departments_raw.replace(";", ",").split(",")
            if code.strip()
        ]

    @property
    def level_codes(self) -> list[str]:
        if not self.levels_raw:
            return []
        return [
            code.strip().upper()
            for code in self.levels_raw.replace(";", ",").split(",")
            if code.strip()
        ]

    @property
    def internal_institutions(self) -> list[str]:
        if not self.institutions_raw:
            return []
        return [
            s.strip()
            for s in self.institutions_raw.replace(";", ",").split(",")
            if s.strip()
        ]

    @property
    def n7_is_included(self) -> bool:
        return any(N7_IDENTIFIER in s.upper() for s in self.internal_institutions)

    @property
    def n7_places(self) -> int | None:
        """Estimate N7's share when multiple institutions split a quota."""
        if self.places is None:
            return None
        institutions = self.internal_institutions
        if not institutions or not self.n7_is_included:
            return self.places
        n = len(institutions)
        if n <= 1:
            return self.places
        # 0 means unlimited — keep as-is
        if self.places == 0:
            return 0
        return max(1, round(self.places / n))


def build_excel_template(
    departments: list[str] | None = None,
    universities: list[str] | None = None,
    frameworks: list[str] | None = None,
    levels: list[str] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accords"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 40

    col_widths = [40, 50, 30, 25, 35, 20, 45, 45]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # University dropdown (column A)
    if universities:
        univ_ws = wb.create_sheet("_Universites")
        for i, name in enumerate(universities, start=1):
            univ_ws.cell(row=i, column=1, value=name)
        univ_ws.sheet_state = "hidden"
        dv_univ = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Universites!$A$1:$A${len(universities)}",
            allow_blank=True,
        )
        dv_univ.sqref = "A2:A10000"
        ws.add_data_validation(dv_univ)

    # Department codes dropdown (column C) — all non-empty subsets
    if departments:
        dept_combos = _all_combinations(departments)
        dept_ws = wb.create_sheet("_Departements")
        for i, combo in enumerate(dept_combos, start=1):
            dept_ws.cell(row=i, column=1, value=combo)
        dept_ws.sheet_state = "hidden"
        dv_dept = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Departements!$A$1:$A${len(dept_combos)}",
            allow_blank=True,
        )
        dv_dept.sqref = "C2:C10000"
        ws.add_data_validation(dv_dept)

    # Level codes dropdown (column D) — all non-empty subsets
    if levels:
        level_combos = _all_combinations(levels)
        level_ws = wb.create_sheet("_Niveaux")
        for i, combo in enumerate(level_combos, start=1):
            level_ws.cell(row=i, column=1, value=combo)
        level_ws.sheet_state = "hidden"
        dv_level = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Niveaux!$A$1:$A${len(level_combos)}",
            allow_blank=True,
        )
        dv_level.sqref = "D2:D10000"
        ws.add_data_validation(dv_level)

    # Framework dropdown (column E)
    if frameworks:
        fw_ws = wb.create_sheet("_Cadres")
        for i, name in enumerate(frameworks, start=1):
            fw_ws.cell(row=i, column=1, value=name)
        fw_ws.sheet_state = "hidden"
        dv_fw = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Cadres!$A$1:$A${len(frameworks)}",
            allow_blank=True,
        )
        dv_fw.sqref = "E2:E10000"
        ws.add_data_validation(dv_fw)

    # Instructions sheet
    info_ws = wb.create_sheet("Instructions")
    info_ws.column_dimensions["A"].width = 80
    instructions = [
        "INSTRUCTIONS IMPORT ACCORDS DE MOBILITE",
        "",
        "- Remplissez les donnees a partir de la ligne 2.",
        "- Etablissement externe : choisissez dans la liste (ou saisissez le nom exact).",
        "- Departements concernes : choisissez dans la liste (toutes les combinaisons sont proposees).",
        "  Laissez vide pour 'tous les departements'.",
        "- Niveaux concernes : choisissez dans la liste. Laissez vide si tous les niveaux.",
        "- Cadre d'accord : choisissez dans la liste ou laissez vide.",
        "- Nombre de places : entier positif, ou 'illimite' si pas de limite.",
        "- Etablissements internes : liste des ecoles partageant le quota (ex: INP-ENSEEIHT;INP-ENSAT).",
        "  Si INP-ENSEEIHT figure dans la liste, le quota N7 est estime automatiquement.",
        "",
        "Les lignes avec des erreurs seront enregistrees dans le journal d'import",
        "pour correction manuelle ulterieure.",
    ]
    for i, line in enumerate(instructions, start=1):
        cell = info_ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _all_combinations(codes: list[str]) -> list[str]:
    """Return all non-empty subsets of codes, sorted by length then alphabetically."""
    sorted_codes = sorted(codes)
    result: list[str] = []
    for r in range(1, len(sorted_codes) + 1):
        for combo in combinations(sorted_codes, r):
            result.append(";".join(combo))
    return result


def parse_excel_file(file_bytes: bytes) -> list[ExcelRow]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    sheet_names = [
        name
        for name in wb.sheetnames
        if not name.startswith("_") and name != "Instructions"
    ]
    if not sheet_names:
        raise ValueError(
            "Le fichier Excel ne contient pas de feuille de donnees valide."
        )

    ws = wb[sheet_names[0]]
    rows: list[ExcelRow] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        raw = {
            "etablissement_externe": _cell_str(
                row, COLUMN_MAP["etablissement_externe"]
            ),
            "nom": _cell_str(row, COLUMN_MAP["nom"]),
            "departements": _cell_str(row, COLUMN_MAP["departements"]),
            "niveaux": _cell_str(row, COLUMN_MAP["niveaux"]),
            "cadre": _cell_str(row, COLUMN_MAP["cadre"]),
            "places": _cell_str(row, COLUMN_MAP["places"]),
            "etablissements_internes": _cell_str(
                row, COLUMN_MAP["etablissements_internes"]
            ),
            "remarques": _cell_str(row, COLUMN_MAP["remarques"]),
        }

        excel_row = ExcelRow(row_number=row_idx, raw=raw)
        excel_row.university_name = raw["etablissement_externe"]
        excel_row.agreement_name = raw["nom"]
        excel_row.departments_raw = raw["departements"]
        excel_row.levels_raw = raw["niveaux"]
        excel_row.framework_raw = raw["cadre"]
        excel_row.institutions_raw = raw["etablissements_internes"]
        excel_row.remarks = raw["remarques"]

        places_raw = raw["places"].lower().strip()
        if places_raw in UNLIMITED_PLACES_VALUES:
            excel_row.places = 0
        else:
            try:
                excel_row.places = max(0, int(float(places_raw)))
            except (ValueError, TypeError):
                excel_row.errors.append(f"Nombre de places invalide: '{raw['places']}'")
                excel_row.places = None

        if not excel_row.university_name:
            excel_row.errors.append("Etablissement externe manquant.")
        if not excel_row.agreement_name:
            excel_row.errors.append("Nom de l'accord manquant.")

        rows.append(excel_row)

    return rows


def _cell_str(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    val = row[index]
    return str(val).strip() if val is not None else ""
