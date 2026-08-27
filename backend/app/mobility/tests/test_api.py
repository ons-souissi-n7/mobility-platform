import io
import json
from datetime import date

import openpyxl
import pytest
from django.test import Client

from app.academic.models import AcademicYear
from app.imports.models import RawImport, RawImportEntity, RawImportStatus
from app.institutions.models import PartnerUniversity
from app.mobility.models import Agreement, AgreementDepartment, AgreementYear
from app.reference.models import Country, CTIRegion, Department


@pytest.mark.django_db
class TestMobilityAgreementAPI:
    def setup_method(self):
        self.client = Client()
        self.country = Country.objects.create(
            iso2="ES",
            name_fr="Espagne",
            name_en="Spain",
            cti_region=CTIRegion.EUROPE_HORS_FRANCE,
        )
        self.university = PartnerUniversity.objects.create(
            moveon_id=1001,
            name="Universidad Test",
            country=self.country,
        )
        self.academic_year = AcademicYear.objects.create(
            label="2026-2027",
            start_date=date(2026, 9, 1),
            end_date=date(2027, 8, 31),
        )
        self.department = Department.objects.create(
            code="SN",
            name="Sciences du Numerique",
        )
        self.agreement = Agreement.objects.create(
            moveon_id="REL-001",
            name="Erasmus outgoing agreement",
            partner_university=self.university,
            direction="outgoing",
            inp_total_places=10,
        )
        self.year = AgreementYear.objects.create(
            agreement=self.agreement,
            academic_year=self.academic_year,
            is_active=True,
            n7_places=4,
        )

    def test_list_agreements(self):
        response = self.client.get("/api/v1/mobility/agreements/")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] >= 1
        assert body["results"][0]["moveon_id"] == "REL-001"

    def test_create_agreement(self):
        payload = {
            "name": "New Erasmus agreement",
            "partner_university_id": self.university.id,
            "direction": "outgoing",
            "inp_total_places": 6,
            "inp_institutions": "N7",
            "remarks": "",
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.post(
            "/api/v1/mobility/agreements/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["name"] == "New Erasmus agreement"

    def test_update_agreement(self):
        payload = {
            "name": "Updated Erasmus agreement",
            "partner_university_id": self.university.id,
            "direction": "both",
            "inp_total_places": 10,
            "inp_institutions": "N7",
            "remarks": "",
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.put(
            f"/api/v1/mobility/agreements/{self.agreement.id}/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert response.json()["name"] == "Updated Erasmus agreement"

    def test_delete_agreement(self):
        agreement = Agreement.objects.create(
            moveon_id="REL-DELETE",
            name="Agreement to delete",
            partner_university=self.university,
        )

        response = self.client.delete(f"/api/v1/mobility/agreements/{agreement.id}/")

        assert response.status_code == 204
        assert not Agreement.objects.filter(pk=agreement.id).exists()

    def test_create_agreement_year(self):
        year2 = AcademicYear.objects.create(
            label="2027-2028",
            start_date=date(2027, 9, 1),
            end_date=date(2028, 8, 31),
        )
        payload = {
            "agreement_id": self.agreement.id,
            "academic_year_id": year2.id,
            "is_active": True,
            "n7_places": 3,
        }

        response = self.client.post(
            "/api/v1/mobility/agreement-years/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["n7_places"] == 3

    def test_create_agreement_year_n7_exceeds_inp_returns_400(self):
        year2 = AcademicYear.objects.create(
            label="2027-2028",
            start_date=date(2027, 9, 1),
            end_date=date(2028, 8, 31),
        )
        payload = {
            "agreement_id": self.agreement.id,
            "academic_year_id": year2.id,
            "is_active": True,
            "n7_places": 99,  # > inp_total_places=10
        }

        response = self.client.post(
            "/api/v1/mobility/agreement-years/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 400

    def test_create_agreement_year_department(self):
        AgreementDepartment.objects.create(
            agreement=self.agreement, department=self.department
        )
        payload = {
            "agreement_year_id": self.year.id,
            "department_id": self.department.id,
            "estimated_places": 2,
        }

        response = self.client.post(
            "/api/v1/mobility/agreement-year-departments/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["estimated_places"] == 2

    def test_list_moveon_import_errors(self):
        RawImport.objects.create(
            source="moveon_partner_university",
            entity=RawImportEntity.PARTNER_UNIVERSITY,
            external_id="3001",
            payload={"country": {"name": "Itallie"}},
            status=RawImportStatus.FAILED,
            error_message="Unknown country name: Itallie",
        )
        RawImport.objects.create(
            source="moveon_agreement",
            entity=RawImportEntity.AGREEMENT,
            external_id="REL-ERR",
            payload={"moveon_id": "REL-ERR", "name": "Broken agreement"},
            status=RawImportStatus.FAILED,
            error_message="partner_university_moveon_id is required",
        )

        response = self.client.get("/api/v1/mobility/raw-imports/moveon-errors/")

        assert response.status_code == 200
        data = response.json()
        assert {item["external_id"] for item in data["results"]} == {"REL-ERR"}
        assert "3001" not in {item["external_id"] for item in data["results"]}

    def test_retry_agreement_import_with_partner_university_correction(self):
        raw_import = RawImport.objects.create(
            source="moveon_agreement",
            entity=RawImportEntity.AGREEMENT,
            external_id="REL-RETRY",
            payload={"moveon_id": "REL-RETRY", "name": "Retry agreement"},
            status=RawImportStatus.FAILED,
            error_message="partner_university_moveon_id is required",
        )

        response = self.client.put(
            f"/api/v1/mobility/raw-imports/{raw_import.id}/retry/",
            data=json.dumps({"partner_university_id": self.university.id}),
            content_type="application/json",
        )

        assert response.status_code == 200
        raw_import.refresh_from_db()
        assert raw_import.status == RawImportStatus.IMPORTED
        assert Agreement.objects.filter(moveon_id="REL-RETRY").exists()

    def test_ignore_raw_import(self):
        raw_import = RawImport.objects.create(
            source="moveon_partner_university",
            entity=RawImportEntity.PARTNER_UNIVERSITY,
            external_id="3001",
            payload={"country": {"name": "Itallie"}},
            status=RawImportStatus.FAILED,
            error_message="Unknown country name: Itallie",
        )

        response = self.client.put(
            f"/api/v1/mobility/raw-imports/{raw_import.id}/ignore/"
        )

        assert response.status_code == 200
        raw_import.refresh_from_db()
        assert raw_import.status == RawImportStatus.IGNORED
        assert "Traité manuellement" in raw_import.error_message

    def test_list_agreements_select_options(self):
        response = self.client.get("/api/v1/mobility/agreements/select-options/")

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert "label" in data[0]
        assert "REL-001" in data[0]["label"] or "Erasmus" in data[0]["label"]

    def test_list_agreements_with_search_filter(self):
        Agreement.objects.create(
            moveon_id="REL-002",
            name="Other agreement",
            partner_university=self.university,
        )

        response = self.client.get("/api/v1/mobility/agreements/?search=erasmus")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["moveon_id"] == "REL-001"

    def test_update_agreement_not_found(self):
        payload = {
            "name": "Non existent",
            "partner_university_id": self.university.id,
            "direction": "outgoing",
            "inp_total_places": 5,
            "inp_institutions": "N7",
            "remarks": "",
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.put(
            "/api/v1/mobility/agreements/99999/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 404

    def test_create_agreement_with_duration_weeks(self):
        payload = {
            "name": "Agreement with duration",
            "partner_university_id": self.university.id,
            "direction": "outgoing",
            "inp_total_places": 6,
            "inp_institutions": "N7",
            "remarks": "",
            "duration_weeks": 24,
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.post(
            "/api/v1/mobility/agreements/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["duration_weeks"] == 24
        assert (
            Agreement.objects.get(name="Agreement with duration").duration_weeks == 24
        )

    def test_create_agreement_without_duration_weeks_defaults_to_null(self):
        payload = {
            "name": "Agreement without duration",
            "partner_university_id": self.university.id,
            "direction": "outgoing",
            "inp_total_places": 6,
            "inp_institutions": "N7",
            "remarks": "",
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.post(
            "/api/v1/mobility/agreements/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["duration_weeks"] is None

    def test_update_agreement_duration_weeks(self):
        payload = {
            "name": self.agreement.name,
            "partner_university_id": self.university.id,
            "direction": "outgoing",
            "inp_total_places": 10,
            "inp_institutions": "N7",
            "remarks": "",
            "duration_weeks": 16,
            "department_ids": [],
            "level_ids": [],
        }

        response = self.client.put(
            f"/api/v1/mobility/agreements/{self.agreement.id}/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert response.json()["duration_weeks"] == 16
        self.agreement.refresh_from_db()
        assert self.agreement.duration_weeks == 16

    def test_create_agreement_year_with_duration_weeks_override(self):
        year2 = AcademicYear.objects.create(
            label="2027-2028",
            start_date=date(2027, 9, 1),
            end_date=date(2028, 8, 31),
        )
        payload = {
            "agreement_id": self.agreement.id,
            "academic_year_id": year2.id,
            "is_active": True,
            "n7_places": 3,
            "duration_weeks": 8,
        }

        response = self.client.post(
            "/api/v1/mobility/agreement-years/",
            data=json.dumps(payload),
            content_type="application/json",
        )

        assert response.status_code == 201
        assert response.json()["duration_weeks"] == 8

    def test_export_agreements_excel_includes_duration_header_and_value(self):
        self.agreement.duration_weeks = 20
        self.agreement.save(update_fields=["duration_weeks"])

        response = self.client.get("/api/v1/mobility/agreements/export-excel/")

        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Accords"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Durée (semaines)" in headers
        duration_col = headers.index("Durée (semaines)")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        matching = [r for r in rows if r[0] == self.agreement.name]
        assert len(matching) == 1
        assert matching[0][duration_col] == 20

    def test_export_agreements_excel_prefers_agreement_year_duration_override(self):
        self.agreement.duration_weeks = 20
        self.agreement.save(update_fields=["duration_weeks"])
        self.year.duration_weeks = 6
        self.year.save(update_fields=["duration_weeks"])

        response = self.client.get(
            f"/api/v1/mobility/agreements/export-excel/?year_label={self.academic_year.label}"
        )

        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Accords"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        duration_col = headers.index("Durée (semaines)")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        matching = [r for r in rows if r[0] == self.agreement.name]
        assert len(matching) == 1
        assert matching[0][duration_col] == 6


@pytest.mark.django_db
class TestAgreementDepartmentAPI:
    def setup_method(self):
        self.client = Client()
        self.country = Country.objects.create(
            iso2="DE",
            name_fr="Allemagne",
            name_en="Germany",
            cti_region=CTIRegion.EUROPE_HORS_FRANCE,
        )
        self.university = PartnerUniversity.objects.create(
            moveon_id=2001,
            name="TU Berlin",
            country=self.country,
        )
        self.department = Department.objects.create(
            code="SN", name="Sciences du Numerique"
        )
        self.agreement = Agreement.objects.create(
            moveon_id="REL-DEPT-001",
            name="Agreement with departments",
            partner_university=self.university,
        )
        from app.mobility.models import AgreementDepartment

        self.ad = AgreementDepartment.objects.create(
            agreement=self.agreement, department=self.department
        )

    def test_list_agreement_departments(self):
        response = self.client.get("/api/v1/mobility/agreement-departments/")

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["department_id"] == self.department.id
        assert data[0]["agreement_id"] == self.agreement.id

    def test_list_agreement_departments_filter_by_agreement(self):
        agreement2 = Agreement.objects.create(
            moveon_id="REL-DEPT-002",
            name="Another agreement",
            partner_university=self.university,
        )
        dept2 = Department.objects.create(code="TC", name="Tronc Commun")
        from app.mobility.models import AgreementDepartment

        AgreementDepartment.objects.create(agreement=agreement2, department=dept2)

        response = self.client.get(
            f"/api/v1/mobility/agreement-departments/?agreement_id={self.agreement.id}"
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["department_id"] == self.department.id
