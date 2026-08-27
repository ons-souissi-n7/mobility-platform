"""Tests for the agreements Excel template/import pipeline — focus on the
"Durée du séjour (semaines)" column (duration_weeks)."""

from __future__ import annotations

import io

import openpyxl
import pytest
from django.test import Client

from app.imports.models import RawImport, RawImportEntity
from app.institutions.models import PartnerUniversity
from app.mobility.models import Agreement
from app.mobility.services.excel_importer import (
    COLUMN_MAP,
    TEMPLATE_COLUMNS,
    build_excel_template,
    parse_excel_file,
)
from app.mobility.services.sync_excel import sync_agreements_from_excel
from app.reference.models import Country, CTIRegion


def make_workbook_bytes(rows: list[list[str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(TEMPLATE_COLUMNS))
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def base_row(**overrides) -> list[str]:
    values = {
        "etablissement_externe": "TU Berlin",
        "nom": "Erasmus TU Berlin",
        "departements": "SN",
        "niveaux": "",
        "cadre": "",
        "places": "5",
        "etablissements_internes": "",
        "remarques": "",
        "duree": "24",
    }
    values.update(overrides)
    # Preserve TEMPLATE_COLUMNS / COLUMN_MAP ordering.
    ordered = sorted(COLUMN_MAP.items(), key=lambda item: item[1])
    return [values[key] for key, _ in ordered]


# ---------------------------------------------------------------------------
# Unit tests — template + parsing
# ---------------------------------------------------------------------------


class TestExcelTemplateDurationColumn:
    def test_template_headers_include_duration_column(self):
        assert "duree" in COLUMN_MAP
        assert TEMPLATE_COLUMNS[COLUMN_MAP["duree"]] == "Duree du sejour (semaines)"

    def test_build_excel_template_writes_duration_header(self):
        data = build_excel_template()
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Accords"]
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Duree du sejour (semaines)" in header_row


class TestParseExcelFileDuration:
    def test_valid_duration_is_parsed_as_int(self):
        rows = parse_excel_file(make_workbook_bytes([base_row(duree="24")]))
        assert rows[0].duration_weeks == 24
        assert rows[0].is_valid

    def test_empty_duration_is_none_and_row_stays_valid(self):
        rows = parse_excel_file(make_workbook_bytes([base_row(duree="")]))
        assert rows[0].duration_weeks is None
        assert rows[0].is_valid

    def test_non_numeric_duration_is_recorded_as_error(self):
        rows = parse_excel_file(make_workbook_bytes([base_row(duree="abc")]))
        assert rows[0].duration_weeks is None
        assert not rows[0].is_valid
        assert any("Duree du sejour" in e for e in rows[0].errors)

    def test_decimal_duration_is_truncated_to_int(self):
        rows = parse_excel_file(make_workbook_bytes([base_row(duree="24.7")]))
        assert rows[0].duration_weeks == 24


# ---------------------------------------------------------------------------
# Integration tests — full sync pipeline writes duration_weeks to the DB
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestSyncAgreementsFromExcelDuration:
    def setup_method(self):
        self.country = Country.objects.create(
            iso2="DE",
            name_fr="Allemagne",
            name_en="Germany",
            cti_region=CTIRegion.EUROPE_HORS_FRANCE,
        )
        # _resolve_or_create_university requires a pre-existing match by
        # name/short_name — it does not create one on the fly.
        self.university = PartnerUniversity.objects.create(
            name="TU Berlin", country=self.country
        )

    def test_creates_agreement_with_duration_weeks(self):
        file_bytes = make_workbook_bytes([base_row(duree="16")])

        result = sync_agreements_from_excel(file_bytes)

        assert result.created == 1
        agreement = Agreement.objects.get(name="Erasmus TU Berlin")
        assert agreement.duration_weeks == 16

    def test_creates_agreement_with_no_duration_leaves_it_null(self):
        file_bytes = make_workbook_bytes([base_row(duree="")])

        sync_agreements_from_excel(file_bytes)

        agreement = Agreement.objects.get(name="Erasmus TU Berlin")
        assert agreement.duration_weeks is None

    def test_reimport_updates_existing_agreement_duration_weeks(self):
        Agreement.objects.create(
            name="Erasmus TU Berlin",
            partner_university=self.university,
            duration_weeks=10,
            remarks="old remark",
        )

        file_bytes = make_workbook_bytes([base_row(duree="20", remarques="new remark")])
        result = sync_agreements_from_excel(file_bytes)

        assert result.updated == 1
        agreement = Agreement.objects.get(name="Erasmus TU Berlin")
        assert agreement.duration_weeks == 20
        assert agreement.remarks == "new remark"

    def test_invalid_duration_row_is_not_imported(self):
        file_bytes = make_workbook_bytes([base_row(duree="not-a-number")])

        result = sync_agreements_from_excel(file_bytes)

        assert result.failed == 1
        assert not Agreement.objects.filter(name="Erasmus TU Berlin").exists()


@pytest.mark.django_db
class TestSyncAgreementsFromExcelNoInformationLoss:
    """Exercises the real processing path end-to-end (exactly what the
    django-q worker calls via run_import_excel_agreements), then checks
    failures surface through the same API the admin error panel uses.
    Catches regressions where a row would be silently dropped instead of
    recorded as an error."""

    def setup_method(self):
        self.client = Client()
        self.country = Country.objects.create(
            iso2="DE",
            name_fr="Allemagne",
            name_en="Germany",
            cti_region=CTIRegion.EUROPE_HORS_FRANCE,
        )
        self.university = PartnerUniversity.objects.create(
            name="TU Berlin", country=self.country
        )

    def test_every_row_is_accounted_for_no_silent_loss(self):
        file_bytes = make_workbook_bytes(
            [
                base_row(nom="Accord Valide", duree="20"),
                base_row(
                    nom="Accord Univ Inconnue",
                    etablissement_externe="Universite Fantome",
                ),
                base_row(nom="Accord Places Invalides", places="pas-un-nombre"),
            ]
        )

        before = RawImport.objects.filter(entity=RawImportEntity.AGREEMENT).count()
        result = sync_agreements_from_excel(file_bytes, source_file="test.xlsx")
        after = RawImport.objects.filter(entity=RawImportEntity.AGREEMENT).count()

        assert result.total == 3
        assert result.created == 1
        assert result.failed == 2
        # One RawImport per input row — success or failure, nothing dropped.
        assert after - before == 3
        assert Agreement.objects.filter(name="Accord Valide").exists()
        assert not Agreement.objects.filter(name="Accord Univ Inconnue").exists()
        assert not Agreement.objects.filter(name="Accord Places Invalides").exists()

    def test_failed_rows_appear_in_the_error_panel_api(self):
        file_bytes = make_workbook_bytes(
            [base_row(nom="Accord Erreur", etablissement_externe="Universite Fantome")]
        )
        sync_agreements_from_excel(file_bytes, source_file="test.xlsx")

        response = self.client.get("/api/v1/mobility/raw-imports/moveon-errors/")

        assert response.status_code == 200
        matching = [
            r
            for r in response.json()["results"]
            if "Accord Erreur" in r["payload"].get("nom", "")
        ]
        assert len(matching) == 1
        assert "Universite Fantome" in matching[0]["error_message"]
        assert "Université introuvable" in matching[0]["error_message"]

    def test_successful_rows_do_not_appear_as_errors(self):
        file_bytes = make_workbook_bytes([base_row(nom="Accord OK")])
        sync_agreements_from_excel(file_bytes, source_file="test.xlsx")

        response = self.client.get("/api/v1/mobility/raw-imports/moveon-errors/")

        matching = [
            r
            for r in response.json()["results"]
            if r["payload"].get("nom") == "Accord OK"
        ]
        assert matching == []
