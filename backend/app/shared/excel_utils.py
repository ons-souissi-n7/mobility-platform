"""Utilitaires partagés pour la génération de fichiers Excel."""

from io import BytesIO

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill

BLUE_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center")


def write_header_row(
    ws, headers: list[str], col_widths: list[int] | None = None
) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    if col_widths:
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    else:
        # Auto-largeur basée sur l'en-tête
        for col, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(
                len(header) + 4, 14
            )


def workbook_response(wb: openpyxl.Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_filename(*parts: str, extension: str = "xlsx") -> str:
    """Construit un nom de fichier propre depuis les parties non vides."""
    slug = "_".join(
        p.replace("/", "-").replace(" ", "_").replace("\\", "-") for p in parts if p
    )
    return f"{slug}.{extension}"
