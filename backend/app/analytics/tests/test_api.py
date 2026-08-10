from datetime import date
from io import BytesIO

import openpyxl
import pytest
from django.test import Client

from app.academic.models import AcademicYear


def _make_year(label: str, *, start_year: int) -> AcademicYear:
    return AcademicYear.objects.create(
        label=label,
        start_date=date(start_year, 9, 1),
        end_date=date(start_year + 1, 8, 31),
    )


@pytest.fixture
def client():
    return Client()


@pytest.mark.django_db
class TestTrendsEndpoint:
    def test_returns_200_with_data_key(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/trends/?years=1")
        assert res.status_code == 200
        body = res.json()
        assert "data" in body
        assert len(body["data"]) == 1
        assert body["data"][0]["label"] == "2024-2025"

    def test_all_fields_present(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/trends/?years=1")
        point = res.json()["data"][0]
        for field in ("label", "outgoing", "incoming", "internships", "complementary"):
            assert field in point

    def test_empty_years_returns_empty_list(self, client):
        res = client.get("/api/v1/analytics/trends/?years=5")
        assert res.status_code == 200
        assert res.json()["data"] == []


@pytest.mark.django_db
class TestBreakdownEndpoints:
    def test_by_department_returns_three_keys(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/by-department/?years=1")
        assert res.status_code == 200
        body = res.json()
        for key in ("outgoing", "incoming", "internships"):
            assert key in body
            assert "years" in body[key]
            assert "series" in body[key]

    def test_by_country_returns_three_keys(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/by-country/?years=1")
        assert res.status_code == 200
        body = res.json()
        for key in ("outgoing", "incoming", "internships"):
            assert key in body

    def test_by_university_returns_three_keys(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/by-university/?years=1")
        assert res.status_code == 200
        body = res.json()
        for key in ("outgoing", "incoming", "internships"):
            assert key in body


@pytest.mark.django_db
class TestExportEndpoint:
    def test_returns_xlsx_content_type(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/export/?years=1")
        assert res.status_code == 200
        assert "spreadsheetml" in res.get("Content-Type", "")

    def test_content_disposition_has_filename(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/export/?years=1")
        cd = res.get("Content-Disposition", "")
        assert "statistiques_mobilite" in cd
        assert ".xlsx" in cd

    def test_response_body_is_valid_xlsx(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/export/?years=1")
        wb = openpyxl.load_workbook(BytesIO(res.content))
        assert "Flux globaux" in wb.sheetnames
        assert "Par département" in wb.sheetnames
        assert "Par pays" in wb.sheetnames
        assert "Par université" in wb.sheetnames

    def test_trends_sheet_has_header_row(self, client):
        _make_year("2024-2025", start_year=2024)
        res = client.get("/api/v1/analytics/export/?years=1")
        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb["Flux globaux"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Année", "Sortants", "Entrants", "Stages", "Complémentaires"]

    def test_export_with_no_years_returns_empty_trends_sheet(self, client):
        res = client.get("/api/v1/analytics/export/?years=3")
        assert res.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb["Flux globaux"]
        assert ws.max_row == 1  # only header row
