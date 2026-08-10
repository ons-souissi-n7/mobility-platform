from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .services import (
    get_country_breakdown,
    get_department_breakdown,
    get_mobility_trends,
    get_university_breakdown,
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEADER_ALIGN = Alignment(horizontal="center")


def _style_header_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center")


def _write_trends(wb: openpyxl.Workbook, n_years: int) -> None:
    ws = wb.active
    ws.title = "Flux globaux"
    ws.append(["Année", "Sortants", "Entrants", "Stages", "Complémentaires"])
    _style_header_row(ws, 1)
    for row in get_mobility_trends(n_years):
        ws.append(
            [
                row["label"],
                row["outgoing"],
                row["incoming"],
                row["internships"],
                row["complementary"],
            ]
        )


def _write_breakdown(wb: openpyxl.Workbook, title: str, data: dict) -> None:
    ws = wb.create_sheet(title=title)
    flow_labels = {
        "outgoing": "Sortants",
        "incoming": "Entrants",
        "internships": "Stages",
    }
    current_row = 1

    for flow_key, flow_label in flow_labels.items():
        bk = data[flow_key]
        if not bk["series"]:
            continue
        ws.append([flow_label] + bk["years"])
        _style_header_row(ws, current_row)
        current_row += 1
        for series in bk["series"]:
            ws.append([series["label"]] + series["values"])
            current_row += 1
        ws.append([])
        current_row += 1


def generate_analytics_excel(n_years: int = 5) -> bytes:
    wb = openpyxl.Workbook()
    _write_trends(wb, n_years)
    _write_breakdown(wb, "Par département", get_department_breakdown(n_years))
    _write_breakdown(wb, "Par pays", get_country_breakdown(n_years))
    _write_breakdown(wb, "Par université", get_university_breakdown(n_years))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
