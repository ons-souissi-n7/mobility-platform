import logging

import openpyxl
import openpyxl.worksheet.datavalidation
from django.db import transaction
from django.db.models import (
    Case,
    CharField,
    Count,
    Exists,
    F,
    IntegerField,
    Max,
    OuterRef,
    Q,
    When,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django_fsm import TransitionNotAllowed
from ninja import File, Query, Router, UploadedFile
from ninja.errors import HttpError
from openpyxl.utils import get_column_letter

from app.academic.api import get_academic_year
from app.academic.models import AcademicYear
from app.audit.logger import log_action
from app.imports.models import (
    ImportReport,
    ImportSource,
    RawImport,
    RawImportEntity,
    RawImportStatus,
)
from app.mobility.models import Agreement, AgreementYear, AgreementYearDepartment
from app.shared.api_helpers import (
    LargePaginationQuery,
    PagedResponse,
    PaginationQuery,
    paginate,
)
from app.shared.excel_utils import (
    build_filename,
    format_agreement_label,
    format_university_label,
    workbook_response,
    write_header_row,
)
from app.shared.validators import DomainValidationError
from app.students.models import AnnualEnrollment, Student, StudentWish
from app.students.schemas import StudentRawImportOut
from app.students.services.adapters import excel_wishes as excel_wishes_adapter
from app.students.services.student_transformer import transform_wish
from app.students.services.student_validator import validate_wish
from app.students.services.sync_moveon import WishRow, import_wish_rows

from .models import (
    Assignment,
    AssignmentResult,
    AssignmentStatus,
    ResultSource,
    SlotType,
)
from .schemas import (
    AgreementWishOut,
    AgreementYearOptionOut,
    AssignmentAgreementDeptStat,
    AssignmentAgreementStat,
    AssignmentOut,
    AssignmentResultOut,
    AssignmentResultOverrideIn,
    AssignmentStatsOut,
    OverrideImportReportOut,
    StudentWishesOut,
    WishImportRetryIn,
    WishPatchIn,
)
from .tasks import enqueue_import_excel_wishes, enqueue_sync_moveon_wishes

logger = logging.getLogger(__name__)

router = Router()


def _derive_override_rank(
    annual_enrollment_id: int, agreement_year_id: int | None
) -> int | None:
    """Retourne le rang du vœu étudiant correspondant à l'accord de correction, ou None si hors vœux."""
    if agreement_year_id is None:
        return None
    wish = StudentWish.objects.filter(
        annual_enrollment_id=annual_enrollment_id,
        agreement_year_id=agreement_year_id,
    ).first()
    return wish.rank if wish else None


def _derive_slot_type(agreement_year_id: int | None, department_id: int) -> SlotType:
    """Calcule le SlotType effectif pour un accord donné et un département étudiant."""
    if agreement_year_id is None:
        return SlotType.UNASSIGNED
    has_dept_quota = (
        AgreementYearDepartment.objects.filter(
            agreement_year_id=agreement_year_id,
            agreement_department__department_id=department_id,
        )
        .filter(
            Q(adjusted_places__gt=0)
            | Q(adjusted_places__isnull=True, estimated_places__gt=0)
        )
        .exists()
    )
    return SlotType.DEPT if has_dept_quota else SlotType.ALTERNATIVE


@router.get(
    "/assignments/",
    response=PagedResponse[AssignmentOut],
    summary="Lister les runs d'affectation",
)
def list_assignments(
    request,
    year_id: int | None = None,
    pagination: PaginationQuery = Query(),
):
    qs = Assignment.objects.select_related("academic_year").order_by("-created_at")
    if year_id:
        qs = qs.filter(academic_year_id=year_id)
    count, items = paginate(qs, pagination.page, pagination.page_size)
    return PagedResponse(
        count=count, page=pagination.page, page_size=pagination.page_size, results=items
    )


@router.get(
    "/assignments/{assignment_id}/",
    response=AssignmentOut,
    summary="Detail d'un run d'affectation",
)
def get_assignment(request, assignment_id: int):
    try:
        return Assignment.objects.select_related("academic_year").get(pk=assignment_id)
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc


@router.get(
    "/assignments/{assignment_id}/results/",
    response=PagedResponse[AssignmentResultOut],
    summary="Resultats d'un run d'affectation",
)
def list_assignment_results(
    request,
    assignment_id: int,
    slot_type: str | None = None,
    department_id: int | None = None,
    search: str | None = None,
    pagination: LargePaginationQuery = Query(),
):
    if not Assignment.objects.filter(pk=assignment_id).exists():
        raise HttpError(404, "Affectation introuvable.")

    qs = (
        AssignmentResult.objects.filter(assignment_id=assignment_id)
        .select_related(
            "annual_enrollment__student",
            "annual_enrollment__department",
            "agreement_year__agreement__partner_university__country",
            "override_agreement_year__agreement__partner_university__country",
        )
        .order_by(
            "annual_enrollment__student__last_name",
            "annual_enrollment__student__first_name",
        )
    )
    if slot_type:
        qs = qs.filter(slot_type=slot_type)
    if department_id:
        qs = qs.filter(annual_enrollment__department_id=department_id)
    if search:
        qs = qs.filter(
            Q(annual_enrollment__student__last_name__icontains=search)
            | Q(annual_enrollment__student__first_name__icontains=search)
            | Q(annual_enrollment__student__ine__icontains=search)
        )
    count, items = paginate(qs, pagination.page, pagination.page_size)
    return PagedResponse(
        count=count, page=pagination.page, page_size=pagination.page_size, results=items
    )


@router.get(
    "/assignments/{assignment_id}/stats/",
    response=AssignmentStatsOut,
    summary="Statistiques d'un run d'affectation",
)
def get_assignment_stats(request, assignment_id: int):
    try:
        assignment = Assignment.objects.get(pk=assignment_id)
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    # Annotations pour les valeurs effectives (après overrides manuels)
    # eff_slot : slot type final (override_slot_type si correction, sinon slot_type auto)
    # eff_ay_id : agreement_year_id final (override_agreement_year_id si correction, sinon auto)
    qs = AssignmentResult.objects.filter(assignment=assignment).annotate(
        eff_slot=Case(
            When(
                source=ResultSource.OVERRIDE,
                then=Coalesce(F("override_slot_type"), F("slot_type")),
            ),
            default=F("slot_type"),
            output_field=CharField(max_length=20),
        ),
        eff_ay_id=Case(
            When(source=ResultSource.OVERRIDE, then=F("override_agreement_year_id")),
            default=F("agreement_year_id"),
            output_field=IntegerField(),
        ),
    )

    slot_counts = dict(
        qs.values("eff_slot").annotate(n=Count("id")).values_list("eff_slot", "n")
    )

    eff_assigned_count = qs.exclude(eff_slot=SlotType.UNASSIGNED).count()
    eff_unassigned_count = qs.filter(eff_slot=SlotType.UNASSIGNED).count()

    dept_counts = list(
        qs.exclude(eff_slot=SlotType.UNASSIGNED)
        .values(
            "annual_enrollment__department_id",
            "annual_enrollment__department__code",
            "annual_enrollment__department__name",
        )
        .annotate(count=Count("id"))
        .order_by("annual_enrollment__department__code")
    )

    # Lignes effectivement affectées avec leur ay effectif
    effective_rows = list(
        qs.exclude(eff_slot=SlotType.UNASSIGNED)
        .exclude(eff_ay_id__isnull=True)
        .values(
            "eff_ay_id",
            "annual_enrollment__department__code",
            "annual_enrollment__department__name",
        )
    )

    ay_ids = list({r["eff_ay_id"] for r in effective_rows})

    # Détails des accord-années effectives (noms, université, pays)
    ay_detail_map = {
        ay.id: ay
        for ay in AgreementYear.objects.filter(pk__in=ay_ids).select_related(
            "agreement__partner_university__country"
        )
    }

    # Comptage affectés par (eff_ay_id × dept)
    assigned_by_ay_dept: dict[int, dict[str, dict]] = {}
    for row in effective_rows:
        ay_id = row["eff_ay_id"]
        code = row["annual_enrollment__department__code"]
        bucket = assigned_by_ay_dept.setdefault(ay_id, {})
        if code in bucket:
            bucket[code]["assigned"] += 1
        else:
            bucket[code] = {
                "assigned": 1,
                "dept_name": row["annual_enrollment__department__name"],
            }

    # Comptage affectés par eff_ay_id
    from collections import Counter

    ay_assigned_count = Counter(r["eff_ay_id"] for r in effective_rows)

    # Quotas réservés par accord × département (AgreementYearDepartment)
    quotas_by_ay_dept: dict[int, dict[str, dict]] = {}
    for row in (
        AgreementYearDepartment.objects.filter(agreement_year_id__in=ay_ids)
        .annotate(
            dept_code=F("agreement_department__department__code"),
            dept_name=F("agreement_department__department__name"),
            effective_quota=Coalesce("adjusted_places", "estimated_places"),
        )
        .values("agreement_year_id", "dept_code", "dept_name", "effective_quota")
    ):
        ay_id = row["agreement_year_id"]
        quotas_by_ay_dept.setdefault(ay_id, {})[row["dept_code"]] = {
            "quota": row["effective_quota"],
            "dept_name": row["dept_name"],
        }

    by_agreement = []
    for ay_id, n_assigned in sorted(ay_assigned_count.items()):
        ay = ay_detail_map.get(ay_id)
        if not ay:
            continue

        dept_map: dict[str, AssignmentAgreementDeptStat] = {}
        for code, q in quotas_by_ay_dept.get(ay_id, {}).items():
            a = assigned_by_ay_dept.get(ay_id, {}).get(code, {})
            dept_map[code] = AssignmentAgreementDeptStat(
                dept_code=code,
                dept_name=q["dept_name"],
                quota=q["quota"],
                assigned=a.get("assigned", 0),
            )
        for code, a in assigned_by_ay_dept.get(ay_id, {}).items():
            if code not in dept_map:
                dept_map[code] = AssignmentAgreementDeptStat(
                    dept_code=code,
                    dept_name=a["dept_name"],
                    quota=0,
                    assigned=a["assigned"],
                )

        by_agreement.append(
            AssignmentAgreementStat(
                agreement_year_id=ay_id,
                agreement_name=ay.agreement.name,
                university_name=ay.agreement.partner_university.name,
                total_places=ay.n7_places,
                assigned=n_assigned,
                fill_rate=round(n_assigned / max(ay.n7_places, 1) * 100, 1),
                by_department=sorted(dept_map.values(), key=lambda x: x.dept_code),
            )
        )

    by_agreement.sort(key=lambda x: x.agreement_name)

    # Comptage par pays — depuis les ay effectives
    country_counter: dict[tuple[str, str], int] = {}
    for ay_id, n in ay_assigned_count.items():
        ay = ay_detail_map.get(ay_id)
        if not ay:
            continue
        country = ay.agreement.partner_university.country
        key = (
            country.name_fr if country else "Inconnu",
            country.iso2 if country else "",
        )
        country_counter[key] = country_counter.get(key, 0) + n

    # Comptage par (dept × pays)
    dept_country_counter: dict[tuple[str, str, str, str], int] = {}
    for row in effective_rows:
        ay_id = row["eff_ay_id"]
        ay = ay_detail_map.get(ay_id)
        if not ay:
            continue
        country = ay.agreement.partner_university.country
        key = (
            row["annual_enrollment__department__code"],
            row["annual_enrollment__department__name"],
            country.name_fr if country else "Inconnu",
            country.iso2 if country else "",
        )
        dept_country_counter[key] = dept_country_counter.get(key, 0) + 1

    # Total inscrits pour l'année (tous les étudiants avec affectation, assignés ou non)
    from app.students.models import AnnualEnrollment

    total_enrolled = AnnualEnrollment.objects.filter(
        academic_year=assignment.academic_year
    ).count()

    # Total par département (tous étudiants, affectés ou non) depuis les résultats d'affectation
    total_by_dept = dict(
        qs.values("annual_enrollment__department__code")
        .annotate(n=Count("id"))
        .values_list("annual_enrollment__department__code", "n")
    )

    return AssignmentStatsOut(
        total_students=assignment.total_students,
        assigned_count=eff_assigned_count,
        unassigned_count=eff_unassigned_count,
        total_enrolled=total_enrolled,
        by_slot_type=slot_counts,
        total_by_dept=total_by_dept,
        by_department=[
            {
                "department_id": r["annual_enrollment__department_id"],
                "department_code": r["annual_enrollment__department__code"],
                "department_name": r["annual_enrollment__department__name"],
                "count": r["count"],
            }
            for r in dept_counts
        ],
        by_agreement=by_agreement,
        by_country=sorted(
            [
                {"country_name_fr": name_fr, "country_iso2": iso2, "count": cnt}
                for (name_fr, iso2), cnt in country_counter.items()
            ],
            key=lambda x: -x["count"],
        ),
        by_dept_country=[
            {
                "department_code": dept_code,
                "department_name": dept_name,
                "country_name_fr": name_fr,
                "country_iso2": iso2,
                "count": cnt,
            }
            for (dept_code, dept_name, name_fr, iso2), cnt in sorted(
                dept_country_counter.items(), key=lambda x: (x[0][0], -x[1])
            )
        ],
    )


@router.get(
    "/assignments/{assignment_id}/agreement-years/",
    response=list[AgreementYearOptionOut],
    summary="Accords disponibles pour la sélection manuelle d'affectation",
)
def list_assignment_agreement_years(request, assignment_id: int):
    try:
        assignment = Assignment.objects.get(pk=assignment_id)
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    ays = (
        AgreementYear.objects.filter(
            academic_year=assignment.academic_year, is_active=True
        )
        .select_related("agreement__partner_university__country")
        .order_by("agreement__name")
    )
    return [
        AgreementYearOptionOut(
            id=ay.id,
            label=format_agreement_label(
                ay.agreement.name,
                ay.agreement.partner_university.name
                if ay.agreement.partner_university_id
                else "",
                ay.agreement.partner_university.country.name_fr
                if ay.agreement.partner_university_id
                and ay.agreement.partner_university.country_id
                else "",
            ),
        )
        for ay in ays
    ]


# ── Vœux sortants ─────────────────────────────────────────────────────────────


@router.get(
    "/wishes/import-errors/",
    response=PagedResponse[StudentRawImportOut],
    summary="Erreurs d'import vœux MoveON",
)
def list_wish_import_errors(
    request, pagination: PaginationQuery = Query(...), year_id: int | None = None
):
    base = RawImport.objects.filter(
        entity=RawImportEntity.STUDENT, source="moveon_student_wishes"
    )
    if year_id is not None:
        base = base.filter(academic_year_id=year_id)
    latest_ids = (
        base.values("external_id")
        .annotate(latest_id=Max("id"))
        .values_list("latest_id", flat=True)
    )
    qs = RawImport.objects.filter(
        id__in=latest_ids,
        status=RawImportStatus.FAILED,
    ).order_by("-created_at")
    count, items = paginate(qs, pagination.page, pagination.page_size)
    return PagedResponse(
        count=count, page=pagination.page, page_size=pagination.page_size, results=items
    )


@router.put(
    "/wishes/import-errors/{raw_import_id}/retry/",
    response=StudentRawImportOut,
    summary="Relancer un import vœu en corrigeant l'étudiant ou l'accord",
)
def retry_wish_import_error(request, raw_import_id: int, payload: WishImportRetryIn):
    try:
        raw_import = RawImport.objects.get(
            pk=raw_import_id,
            entity=RawImportEntity.STUDENT,
            status=RawImportStatus.FAILED,
        )
    except RawImport.DoesNotExist as exc:
        raise HttpError(404, "Erreur d'import vœu introuvable.") from exc

    if raw_import.academic_year_id is None:
        raise HttpError(400, "Cet import n'est pas associé à une année universitaire.")

    corrected = dict(raw_import.payload)

    if payload.student_id is not None:
        student = Student.objects.filter(pk=payload.student_id).first()
        if student is None:
            raise HttpError(400, f"Étudiant {payload.student_id} introuvable.")
        corrected["ine"] = student.ine

    if payload.agreement_id is not None:
        agreement = Agreement.objects.filter(pk=payload.agreement_id).first()
        if agreement is None:
            raise HttpError(400, f"Accord {payload.agreement_id} introuvable.")
        corrected["offre_de_sejour"] = agreement.name

    if payload.rank is not None:
        corrected["rank"] = payload.rank

    from app.academic.models import AcademicYear as AcademicYearModel

    academic_year = AcademicYearModel.objects.get(pk=raw_import.academic_year_id)

    try:
        tw = transform_wish(corrected)
        validate_wish(tw)
    except (ValueError, DomainValidationError) as exc:
        raise HttpError(400, str(exc)) from exc

    row = WishRow(
        individu=tw.individu,
        offre_de_sejour=tw.offre_de_sejour,
        rank=tw.rank,
        ine=tw.ine,
    )

    report = import_wish_rows([row], academic_year)
    if report.errors or report.unresolved:
        reason = (report.errors or [str(report.unresolved[0])])[0]
        raw_import.payload = corrected
        raw_import.error_message = reason
        raw_import.save(update_fields=["payload", "error_message", "updated_at"])
        raise HttpError(400, reason)

    raw_import.payload = corrected
    raw_import.status = RawImportStatus.IMPORTED
    raw_import.error_message = ""
    from django.utils import timezone as tz

    raw_import.imported_at = tz.now()
    raw_import.save(
        update_fields=[
            "payload",
            "status",
            "error_message",
            "imported_at",
            "updated_at",
        ]
    )
    log_action(
        request,
        action="retry_wish_import_error",
        detail=f"RawImport {raw_import_id} relancé avec succès.",
    )
    return raw_import


@router.get(
    "/wishes/template/{year_id}/",
    summary="Télécharger le template Excel vœux pour une année",
)
def download_wish_template(request, year_id: int):
    academic_year = get_academic_year(year_id)
    file_bytes = excel_wishes_adapter.generate_wish_template(academic_year)
    label_slug = academic_year.label.replace("/", "-")
    response = HttpResponse(
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="template_voeux_{label_slug}.xlsx"'
    )
    return response


@router.post(
    "/wishes/import-excel/{year_id}/",
    response={202: dict},
    summary="Importer les vœux depuis un fichier Excel",
)
def import_wishes_from_excel(request, year_id: int, file: UploadedFile = File(...)):
    academic_year = get_academic_year(year_id)
    triggered_by = getattr(request.user, "username", "")
    file_bytes = file.read()
    task_id = enqueue_import_excel_wishes(
        file_bytes, file.name, year_id, triggered_by=triggered_by
    )
    log_action(
        request,
        action="import_excel_wishes",
        detail=f"Tâche {task_id} lancée — Fichier {file.name} — Année {academic_year.label}",
    )
    return 202, {
        "task_id": task_id,
        "message": "Import Excel vœux lancé en arrière-plan.",
    }


@router.get(
    "/wishes/export-excel/{year_id}/",
    summary="Exporter les vœux en Excel",
)
def export_wishes_excel(
    request,
    year_id: int,
    dept_code: str | None = None,
):
    academic_year = get_academic_year(year_id)

    # Dernière affectation calculée pour cette année (peut être None)
    assignment = (
        Assignment.objects.filter(academic_year=academic_year)
        .order_by("-created_at")
        .first()
    )

    # Lookup enrollment_id → résultat d'affectation
    result_lookup: dict[int, AssignmentResult] = {}
    if assignment:
        for res in AssignmentResult.objects.filter(
            assignment=assignment
        ).select_related(
            "agreement_year__agreement__partner_university__country",
            "override_agreement_year__agreement__partner_university__country",
        ):
            result_lookup[res.annual_enrollment_id] = res

    qs = (
        StudentWish.objects.filter(annual_enrollment__academic_year=academic_year)
        .select_related(
            "annual_enrollment__student__nationality",
            "annual_enrollment__department",
            "annual_enrollment__parcours",
            "annual_enrollment__level",
            "agreement_year__agreement__partner_university__country",
        )
        .order_by(
            "annual_enrollment__student__last_name",
            "annual_enrollment__student__first_name",
            "rank",
        )
    )
    if dept_code:
        qs = qs.filter(annual_enrollment__department__code=dept_code)

    slot_labels = {
        "dept": "Quota département",
        "surplus": "Quota mutualisé",
        "alternative": "Alternative",
    }

    rows: dict[int, dict] = {}
    for w in qs:
        enr = w.annual_enrollment
        sid = enr.student_id
        if sid not in rows:
            res = result_lookup.get(enr.id)
            auto_decision = correction_accord = correction_motif = ""
            assigned_slot = assigned_rank = ""
            if res:
                if res.agreement_year_id:
                    ag = res.agreement_year.agreement
                    univ = ag.partner_university if ag.partner_university_id else None
                    parts = [
                        p
                        for p in [
                            ag.name,
                            univ.name if univ else "",
                            univ.country.name_fr if univ and univ.country_id else "",
                        ]
                        if p
                    ]
                    auto_decision = " — ".join(parts)
                else:
                    auto_decision = "Non affecté"
                assigned_slot = slot_labels.get(res.slot_type, res.slot_type)
                assigned_rank = (
                    res.assigned_rank if res.assigned_rank is not None else ""
                )
                if res.source == ResultSource.OVERRIDE:
                    if res.override_agreement_year_id:
                        _ov_ag = res.override_agreement_year.agreement
                        _ov_univ = (
                            _ov_ag.partner_university
                            if _ov_ag.partner_university_id
                            else None
                        )
                        correction_accord = format_agreement_label(
                            _ov_ag.name,
                            _ov_univ.name if _ov_univ else "",
                            _ov_univ.country.name_fr
                            if _ov_univ and _ov_univ.country_id
                            else "",
                        )
                    correction_motif = res.override_reason or ""
            rows[sid] = {
                "ine": enr.student.ine,
                "nom": enr.student.last_name,
                "prenom": enr.student.first_name,
                "dept": enr.department.code if enr.department_id else "",
                "filiere": enr.parcours.code if enr.parcours_id else "",
                "niveau": enr.level.code if enr.level_id else "",
                "gpa": float(enr.gpa) if enr.gpa is not None else "",
                "nationalite": enr.student.nationality.name_fr
                if enr.student.nationality_id
                else "",
                "auto_decision": auto_decision,
                "assigned_slot": assigned_slot,
                "assigned_rank": assigned_rank,
                "correction_accord": correction_accord,
                "correction_motif": correction_motif,
                "wishes": [],
            }
        agreement = w.agreement_year.agreement
        univ = agreement.partner_university if agreement.partner_university_id else None
        rows[sid]["wishes"].append(
            {
                "accord": agreement.name,
                "universite": univ.name if univ else "",
                "pays": univ.country.name_fr if univ and univ.country_id else "",
                "rank": w.rank,
            }
        )

    max_rank = max(
        (max((ww["rank"] for ww in r["wishes"]), default=0) for r in rows.values()),
        default=3,
    )
    max_rank = max(max_rank, 3)

    label_slug = academic_year.label.replace("/", "-")
    filename = build_filename("mobilite-sortante", label_slug, dept_code or "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mobilité sortante"

    headers = [
        "INE",
        "Nom",
        "Prénom",
        "Département",
        "Filière",
        "Niveau",
        "GPA",
        "Nationalité",
    ]
    widths = [14, 22, 20, 14, 14, 10, 8, 22]
    for r in range(1, max_rank + 1):
        headers.append(f"Vœu {r}")
        widths.append(70)
    if assignment:
        headers += [
            "Auto-affectation",
            "Type de quota",
            "Rang d'affectation",
            "Correction (accord)",
            "Motif de correction",
        ]
        widths += [80, 22, 10, 60, 50]
    write_header_row(ws, headers, widths)

    for r in sorted(rows.values(), key=lambda x: (x["nom"], x["prenom"])):
        wish_map = {w["rank"]: w for w in r["wishes"]}
        row_data = [
            r["ine"],
            r["nom"],
            r["prenom"],
            r["dept"],
            r["filiere"],
            r["niveau"],
            r["gpa"],
            r["nationalite"],
        ]
        for rank in range(1, max_rank + 1):
            w = wish_map.get(rank)
            if w:
                parts = [p for p in [w["accord"], w["universite"], w["pays"]] if p]
                row_data.append(" — ".join(parts))
            else:
                row_data.append("")
        if assignment:
            row_data += [
                r["auto_decision"],
                r["assigned_slot"],
                r["assigned_rank"],
                r["correction_accord"],
                r["correction_motif"],
            ]
        ws.append(row_data)

    # Feuille 2 : accords disponibles + dropdown sur la colonne Correction
    if assignment:
        active_ays = list(
            AgreementYear.objects.filter(academic_year=academic_year, is_active=True)
            .select_related("agreement__partner_university__country")
            .order_by("agreement__name")
        )
        accord_names = [
            format_agreement_label(
                ay.agreement.name,
                ay.agreement.partner_university.name
                if ay.agreement.partner_university_id
                else "",
                ay.agreement.partner_university.country.name_fr
                if ay.agreement.partner_university_id
                and ay.agreement.partner_university.country_id
                else "",
            )
            for ay in active_ays
        ]

        # Feuille cachée pour la validation de données
        if accord_names:
            acc_ws = wb.create_sheet("_Accords")
            # Première option : désaffectation manuelle
            acc_ws.cell(row=1, column=1, value="Aucune affectation")
            for i, label in enumerate(accord_names, start=2):
                acc_ws.cell(row=i, column=1, value=label)
            acc_ws.sheet_state = "hidden"

            total_rows = len(accord_names) + 1
            corr_col_letter = get_column_letter(
                headers.index("Correction (accord)") + 1
            )
            dv = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1=f"_Accords!$A$1:$A${total_rows}",
                allow_blank=True,
            )
            dv.sqref = f"{corr_col_letter}2:{corr_col_letter}10000"
            ws.add_data_validation(dv)

        # Feuille 2 visible : accords disponibles (aide visuelle)
        ws2 = wb.create_sheet("Accords disponibles")
        write_header_row(
            ws2, ["Accord", "Université", "Pays", "Places N7"], [50, 50, 25, 10]
        )
        for ay in active_ays:
            ag = ay.agreement
            univ = ag.partner_university if ag.partner_university_id else None
            ws2.append(
                [
                    ag.name,
                    univ.name if univ else "",
                    univ.country.name_fr if univ and univ.country_id else "",
                    ay.n7_places,
                ]
            )

    return workbook_response(wb, filename)


@router.post(
    "/assignments/{assignment_id}/validate/",
    response={200: AssignmentOut},
    summary="Valider les affectations (proposé → validé)",
)
def validate_assignment(request, assignment_id: int):
    try:
        assignment = Assignment.objects.select_related("academic_year").get(
            pk=assignment_id
        )
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    try:
        assignment.validate()
    except TransitionNotAllowed as exc:
        raise HttpError(
            409, f"Transition impossible depuis l'état '{assignment.status}'."
        ) from exc

    assignment.save(update_fields=["status", "updated_at"])

    log_action(
        request,
        action="validate_assignment",
        detail=f"Affectation {assignment_id} ({assignment.academic_year.label}) validée.",
    )
    return assignment


@router.post(
    "/assignments/{assignment_id}/publish/",
    response={200: AssignmentOut},
    summary="Publier les résultats d'affectation (validé → publié)",
)
def publish_assignment(request, assignment_id: int):
    try:
        assignment = Assignment.objects.select_related("academic_year").get(
            pk=assignment_id
        )
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    try:
        assignment.publish()
    except TransitionNotAllowed as exc:
        raise HttpError(
            409, f"Transition impossible depuis l'état '{assignment.status}'."
        ) from exc

    assignment.save(update_fields=["status", "updated_at"])

    # Transition année : validation → published
    academic_year = assignment.academic_year
    if academic_year.status == AcademicYear.CampaignStatus.VALIDATION:
        try:
            academic_year.publish_results()
            academic_year.save(update_fields=["status", "updated_at"])
        except TransitionNotAllowed:
            logger.warning(
                "publish_results transition refused for year %s (status=%s)",
                academic_year.label,
                academic_year.status,
            )

    log_action(
        request,
        action="publish_assignment",
        detail=f"Affectation {assignment_id} ({academic_year.label}) publiée.",
    )
    return assignment


@router.post(
    "/assignments/{assignment_id}/import-overrides/",
    response={200: OverrideImportReportOut},
    summary="Importer les corrections de supervision depuis le fichier Excel",
)
def import_assignment_overrides(
    request, assignment_id: int, file: UploadedFile = File(...)
):
    try:
        assignment = Assignment.objects.select_related("academic_year").get(
            pk=assignment_id
        )
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    if assignment.status in (
        AssignmentStatus.VALIDATED,
        AssignmentStatus.PUBLISHED,
        AssignmentStatus.CANCELLED,
    ):
        raise HttpError(
            403,
            f"Modifications interdites — affectation en état '{assignment.status}'.",
        )

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise HttpError(400, "Fichier Excel invalide ou corrompu.") from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise HttpError(400, "Fichier Excel vide.")
    headers = [str(h or "").strip() for h in header_row]

    required_cols = ["INE", "Correction (accord)", "Motif de correction"]
    missing_cols = [col for col in required_cols if col not in headers]
    if missing_cols:
        raise HttpError(
            400,
            f"Colonnes requises manquantes : {', '.join(missing_cols)}",
        )
    ine_col = headers.index("INE")
    correction_col = headers.index("Correction (accord)")
    motif_col = headers.index("Motif de correction")

    ay_by_name: dict[str, int] = {}
    for _ay in AgreementYear.objects.filter(
        academic_year=assignment.academic_year, is_active=True
    ).select_related("agreement__partner_university__country"):
        _ag = _ay.agreement
        _univ = _ag.partner_university if _ag.partner_university_id else None
        _full_label = format_agreement_label(
            _ag.name,
            _univ.name if _univ else "",
            _univ.country.name_fr if _univ and _univ.country_id else "",
        )
        ay_by_name[_full_label.strip().lower()] = _ay.id
        ay_by_name[_ag.name.strip().lower()] = _ay.id

    result_by_ine: dict[str, AssignmentResult] = {
        res.annual_enrollment.student.ine: res
        for res in AssignmentResult.objects.filter(
            assignment=assignment
        ).select_related("annual_enrollment__student")
    }

    updated = 0
    skipped = 0
    errors: list[dict] = []

    with transaction.atomic():
        for row_idx, row in enumerate(rows_iter, start=2):
            ine = str(row[ine_col] or "").strip()
            if not ine:
                continue

            correction = str(row[correction_col] or "").strip()
            motif = str(row[motif_col] or "").strip()

            if not correction:
                skipped += 1
                continue

            if not motif:
                errors.append(
                    {
                        "ine": ine,
                        "ligne": row_idx,
                        "erreur": "Motif de correction obligatoire.",
                    }
                )
                continue

            res = result_by_ine.get(ine)
            if res is None:
                errors.append(
                    {
                        "ine": ine,
                        "ligne": row_idx,
                        "erreur": f"Étudiant INE '{ine}' non trouvé dans cette affectation.",
                    }
                )
                continue

            # Mot-clé spécial : retrait d'affectation → non affecté
            if correction.strip().lower() in ("non affecté", "aucune affectation"):
                res.override_slot_type = SlotType.UNASSIGNED
                res.override_agreement_year_id = None
                res.override_rank = None
                res.override_reason = motif
                res.source = ResultSource.OVERRIDE
                res.save(
                    update_fields=[
                        "override_slot_type",
                        "override_agreement_year_id",
                        "override_rank",
                        "override_reason",
                        "source",
                        "updated_at",
                    ]
                )
                updated += 1
                continue

            ay_id = ay_by_name.get(correction.strip().lower())
            if ay_id is None:
                errors.append(
                    {
                        "ine": ine,
                        "ligne": row_idx,
                        "erreur": f"Accord '{correction}' introuvable pour cette année.",
                    }
                )
                continue

            dept_id = res.annual_enrollment.department_id
            res.override_slot_type = _derive_slot_type(ay_id, dept_id)
            res.override_agreement_year_id = ay_id
            res.override_rank = _derive_override_rank(res.annual_enrollment_id, ay_id)
            res.override_reason = motif
            res.source = ResultSource.OVERRIDE
            res.save(
                update_fields=[
                    "override_slot_type",
                    "override_agreement_year_id",
                    "override_rank",
                    "override_reason",
                    "source",
                    "updated_at",
                ]
            )
            updated += 1

    triggered_by = getattr(request.user, "username", "") if request else ""
    ImportReport.objects.create(
        source=ImportSource.EXCEL_OVERRIDES,
        academic_year=assignment.academic_year,
        triggered_by=triggered_by,
        total=updated + skipped + len(errors),
        success_count=updated,
        error_count=len(errors),
        errors=[
            {"external_id": e["ine"], "reason": f"Ligne {e['ligne']}: {e['erreur']}"}
            for e in errors
        ],
    )

    log_action(
        request,
        action="import_assignment_overrides",
        detail=f"Affectation {assignment_id} — {updated} correction(s) importée(s), {skipped} ignorée(s), {len(errors)} erreur(s).",
    )
    return OverrideImportReportOut(updated=updated, skipped=skipped, errors=errors)


@router.patch(
    "/assignments/{assignment_id}/results/{result_id}/",
    response={200: AssignmentResultOut},
    summary="Modifier individuellement l'affectation d'un étudiant",
)
def override_assignment_result(
    request,
    assignment_id: int,
    result_id: int,
    payload: AssignmentResultOverrideIn,
):
    try:
        assignment = Assignment.objects.get(pk=assignment_id)
    except Assignment.DoesNotExist as exc:
        raise HttpError(404, "Affectation introuvable.") from exc

    if assignment.status in (
        AssignmentStatus.VALIDATED,
        AssignmentStatus.PUBLISHED,
        AssignmentStatus.CANCELLED,
    ):
        raise HttpError(
            403,
            f"Modifications interdites — affectation en état '{assignment.status}'.",
        )

    try:
        result = AssignmentResult.objects.select_related(
            "annual_enrollment__student",
            "annual_enrollment__department",
            "agreement_year__agreement__partner_university",
            "override_agreement_year__agreement",
        ).get(pk=result_id, assignment_id=assignment_id)
    except AssignmentResult.DoesNotExist as exc:
        raise HttpError(404, "Résultat d'affectation introuvable.") from exc

    if payload.force_unassigned:
        if not (payload.override_reason or "").strip():
            raise HttpError(400, "Le motif de désaffectation est obligatoire.")
        result.override_slot_type = SlotType.UNASSIGNED
        result.override_agreement_year_id = None
        result.override_rank = None
        result.override_reason = payload.override_reason
        result.source = ResultSource.OVERRIDE
        result.save(
            update_fields=[
                "override_slot_type",
                "override_agreement_year_id",
                "override_rank",
                "override_reason",
                "source",
                "updated_at",
            ]
        )
    elif payload.override_agreement_year_id is not None:
        if not (payload.override_reason or "").strip():
            raise HttpError(400, "Le motif de correction est obligatoire.")
        dept_id = result.annual_enrollment.department_id
        result.override_slot_type = _derive_slot_type(
            payload.override_agreement_year_id, dept_id
        )
        result.override_agreement_year_id = payload.override_agreement_year_id
        result.override_rank = _derive_override_rank(
            result.annual_enrollment_id, payload.override_agreement_year_id
        )
        result.override_reason = payload.override_reason
        result.source = ResultSource.OVERRIDE
        result.save(
            update_fields=[
                "override_slot_type",
                "override_agreement_year_id",
                "override_rank",
                "override_reason",
                "source",
                "updated_at",
            ]
        )
    else:
        result.override_slot_type = None
        result.override_agreement_year_id = None
        result.override_rank = None
        result.override_reason = payload.override_reason
        result.source = ResultSource.AUTO
        result.save(
            update_fields=[
                "override_slot_type",
                "override_agreement_year_id",
                "override_rank",
                "override_reason",
                "source",
                "updated_at",
            ]
        )

    log_action(
        request,
        action="override_assignment_result",
        detail=(
            f"Résultat {result_id} (étudiant {result.annual_enrollment.student.ine}) modifié."
        ),
    )
    return result


@router.post(
    "/wishes/sync-moveon/{year_id}/",
    response={202: dict},
    summary="Synchroniser les vœux étudiants depuis MoveON",
)
def sync_wishes_from_moveon(request, year_id: int):
    academic_year = get_academic_year(year_id)
    triggered_by = getattr(request.user, "username", "")
    task_id = enqueue_sync_moveon_wishes(year_id, triggered_by=triggered_by)
    log_action(
        request,
        action="sync_moveon_wishes",
        detail=f"Tâche {task_id} lancée — Année {academic_year.label}",
    )
    return 202, {
        "task_id": task_id,
        "message": "Synchronisation MoveON lancée en arrière-plan.",
    }


@router.get(
    "/wishes/by-year/{year_id}/",
    response=PagedResponse[StudentWishesOut],
    summary="Vœux ordonnés par étudiant pour une année",
)
def list_wishes_by_year(request, year_id: int, pagination: PaginationQuery = Query()):
    academic_year = get_academic_year(year_id)

    # Paginate on enrolled students who have at least one wish
    enrollment_qs = (
        AnnualEnrollment.objects.filter(
            academic_year=academic_year,
        )
        .filter(Exists(StudentWish.objects.filter(annual_enrollment_id=OuterRef("pk"))))
        .select_related(
            "student",
            "student__nationality",
            "department",
            "parcours",
            "level",
        )
        .order_by("student__last_name", "student__first_name")
    )
    count, enrollments = paginate(enrollment_qs, pagination.page, pagination.page_size)

    if not enrollments:
        return PagedResponse(
            count=count,
            page=pagination.page,
            page_size=pagination.page_size,
            results=[],
        )

    enrollment_ids = [e.id for e in enrollments]
    enrollment_by_id = {e.id: e for e in enrollments}

    wishes_qs = (
        StudentWish.objects.filter(annual_enrollment_id__in=enrollment_ids)
        .select_related(
            "agreement_year__agreement__partner_university__country",
        )
        .order_by(
            "annual_enrollment__student__last_name",
            "annual_enrollment__student__first_name",
            "rank",
        )
    )

    grouped: dict[int, StudentWishesOut] = {}
    for w in wishes_qs:
        enrollment = enrollment_by_id[w.annual_enrollment_id]
        student = enrollment.student
        sid = student.id
        if sid not in grouped:
            grouped[sid] = StudentWishesOut(
                enrollment_id=enrollment.id,
                student_id=sid,
                ine=student.ine,
                first_name=student.first_name,
                last_name=student.last_name,
                department_code=enrollment.department.code
                if enrollment.department_id
                else None,
                parcours_code=enrollment.parcours.code
                if enrollment.parcours_id
                else None,
                level_code=enrollment.level.code if enrollment.level_id else None,
                gpa=enrollment.gpa,
                nationality_name_fr=student.nationality.name_fr
                if student.nationality_id
                else None,
                is_alternant=enrollment.is_alternant,
                is_scholarship=enrollment.is_scholarship,
                wishes=[],
            )
        agreement = w.agreement_year.agreement
        univ = agreement.partner_university
        grouped[sid].wishes.append(
            AgreementWishOut(
                wish_id=w.id,
                rank=w.rank,
                agreement_id=agreement.id,
                moveon_id=agreement.moveon_id,
                agreement_name=agreement.name,
                university_name=format_university_label(
                    univ.name,
                    univ.short_name or "",
                    univ.country.name_fr if univ.country_id else "",
                ),
                direction=agreement.direction,
            )
        )

    return PagedResponse(
        count=count,
        page=pagination.page,
        page_size=pagination.page_size,
        results=list(grouped.values()),
    )


@router.get(
    "/wishes/by-student/{student_id}/{year_id}/",
    response=StudentWishesOut,
    summary="Vœux d'un étudiant pour une année",
)
def get_student_wishes(request, student_id: int, year_id: int):
    try:
        student = Student.objects.select_related("nationality").get(pk=student_id)
    except Student.DoesNotExist as exc:
        raise HttpError(404, "Étudiant introuvable.") from exc

    academic_year = get_academic_year(year_id)

    enrollment = (
        AnnualEnrollment.objects.filter(student=student, academic_year=academic_year)
        .select_related("department", "parcours")
        .first()
    )

    wishes = (
        (
            StudentWish.objects.filter(annual_enrollment=enrollment)
            .select_related(
                "agreement_year__agreement__partner_university__country",
            )
            .order_by("rank")
        )
        if enrollment
        else StudentWish.objects.none()
    )

    def _wish_out(w) -> AgreementWishOut:
        ag = w.agreement_year.agreement
        univ = ag.partner_university
        return AgreementWishOut(
            wish_id=w.id,
            rank=w.rank,
            agreement_id=ag.id,
            moveon_id=ag.moveon_id,
            agreement_name=ag.name,
            university_name=format_university_label(
                univ.name,
                univ.short_name or "",
                univ.country.name_fr if univ.country_id else "",
            ),
            direction=ag.direction,
        )

    return StudentWishesOut(
        enrollment_id=enrollment.id if enrollment else 0,
        student_id=student.id,
        ine=student.ine,
        first_name=student.first_name,
        last_name=student.last_name,
        department_code=enrollment.department.code if enrollment else None,
        parcours_code=enrollment.parcours.code
        if enrollment and enrollment.parcours_id
        else None,
        level_code=enrollment.level.code
        if enrollment and enrollment.level_id
        else None,
        gpa=enrollment.gpa if enrollment else None,
        nationality_name_fr=student.nationality.name_fr
        if student.nationality_id
        else None,
        is_alternant=enrollment.is_alternant if enrollment else False,
        is_scholarship=enrollment.is_scholarship if enrollment else False,
        wishes=[_wish_out(w) for w in wishes],
    )


@router.delete(
    "/wishes/{wish_id}/",
    response={204: None},
    summary="Supprimer un vœu étudiant",
)
def delete_student_wish(request, wish_id: int):
    try:
        wish = StudentWish.objects.select_related(
            "annual_enrollment__student",
            "annual_enrollment__academic_year",
        ).get(pk=wish_id)
    except StudentWish.DoesNotExist as exc:
        raise HttpError(404, "Vœu introuvable.") from exc
    log_action(
        request,
        action="delete_student_wish",
        detail=(
            f"Vœu #{wish_id} rang {wish.rank} — "
            f"{wish.annual_enrollment.student.ine} — "
            f"Année {wish.annual_enrollment.academic_year.label}"
        ),
    )
    wish.delete()
    return 204, None


@router.patch(
    "/wishes/{wish_id}/",
    response=AgreementWishOut,
    summary="Modifier le rang d'un vœu étudiant",
)
def update_student_wish(request, wish_id: int, payload: WishPatchIn):
    try:
        wish = StudentWish.objects.select_related(
            "agreement_year__agreement__partner_university__country",
            "annual_enrollment__student",
            "annual_enrollment__academic_year",
        ).get(pk=wish_id)
    except StudentWish.DoesNotExist as exc:
        raise HttpError(404, "Vœu introuvable.") from exc
    wish.rank = payload.rank
    wish.save(update_fields=["rank"])
    log_action(
        request,
        action="update_student_wish",
        detail=(
            f"Vœu #{wish_id} rang → {payload.rank} — "
            f"{wish.annual_enrollment.student.ine} — "
            f"Année {wish.annual_enrollment.academic_year.label}"
        ),
    )
    ag = wish.agreement_year.agreement
    univ = ag.partner_university
    return AgreementWishOut(
        wish_id=wish.id,
        rank=wish.rank,
        agreement_id=ag.id,
        moveon_id=ag.moveon_id,
        agreement_name=ag.name,
        university_name=format_university_label(
            univ.name,
            univ.short_name or "",
            univ.country.name_fr if univ.country_id else "",
        ),
        direction=ag.direction,
    )
