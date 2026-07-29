"""Génération du rapport CTI Excel (format VII.D3) — par promotion (niveaux terminaux)."""

from __future__ import annotations

import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.academic.models import AcademicYear

# ─── Constantes durée ─────────────────────────────────────────────────────────

_WEEKS_PER_MONTH = 4.33
_EXCHANGE_1SEM_WEEKS = 20  # ≥ 20 sem. = 1 semestre
_INTERNSHIP_3M_WEEKS = 13  # < 13 sem. ≈ < 3 mois
_INTERNSHIP_6M_WEEKS = 26  # ≥ 26 sem. ≈ > 6 mois


# ─── Styles openpyxl ──────────────────────────────────────────────────────────

_TITLE_FILL = PatternFill("solid", fgColor="1F3864")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=9)
_COL_HDR_FILL = PatternFill("solid", fgColor="BDD7EE")
_COL_HDR_FONT = Font(bold=True, size=9)
_TOTAL_FILL = PatternFill("solid", fgColor="D9D9D9")
_TOTAL_FONT = Font(bold=True, size=9)
_NOTE_FONT = Font(italic=True, color="595959", size=9)
_DATA_FONT = Font(size=9)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── Helpers durée ────────────────────────────────────────────────────────────


def _exchange_cat(weeks: int | None) -> str:
    if not weeks:
        return "lt"
    if weeks < _EXCHANGE_1SEM_WEEKS:
        return "lt"
    if weeks == _EXCHANGE_1SEM_WEEKS:
        return "eq"
    return "gt"


def _internship_cat(weeks: int | None) -> str:
    if not weeks:
        return "lt"
    if weeks < _INTERNSHIP_3M_WEEKS:
        return "lt"
    if weeks < _INTERNSHIP_6M_WEEKS:
        return "mid"
    return "gt"


def _parse_incoming_duration(duration: str) -> str:
    s = (duration or "").lower().strip()
    if re.search(r"2\s*sem|1\s*an|un\s*an|ann[ée]e|12\s*mois", s):
        return "gt"
    if re.search(r"1\s*sem|un\s*sem", s):
        return "eq"
    m = re.search(r"(\d+)\s*mois", s)
    if m:
        weeks = int(m.group(1)) * _WEEKS_PER_MONTH
        if weeks < _EXCHANGE_1SEM_WEEKS:
            return "lt"
        if weeks < _EXCHANGE_1SEM_WEEKS * 2:
            return "eq"
        return "gt"
    return "lt"


def _incoming_gender(civility: str) -> str:
    c = (civility or "").lower()
    return "F" if ("mme" in c or c.startswith("f")) else "M"


def _weeks_to_months(weeks: float) -> float:
    return round(weeks / _WEEKS_PER_MONTH, 1)


def _months_label(weeks: float) -> str:
    """Ex: 20 sem. → '4.6 mois  (~1 semestre)'."""
    if not weeks:
        return "—"
    m = weeks / _WEEKS_PER_MONTH
    if m < 3:
        cat = "< 3 mois"
    elif m < 6:
        cat = "3–6 mois"
    else:
        cat = "> 6 mois"
    return f"{m:.1f} mois  ({cat})"


# ─── Collecte par cohorte ─────────────────────────────────────────────────────


def _get_terminal_enrolled(academic_year: AcademicYear) -> list:
    """Étudiants inscrits en année N à un niveau terminal (is_terminal=True)."""
    from app.reference.models import Level
    from app.students.models import AnnualEnrollment

    terminal_ids = list(
        Level.objects.filter(is_terminal=True).values_list("id", flat=True)
    )
    return list(
        AnnualEnrollment.objects.filter(
            academic_year=academic_year,
            level_id__in=terminal_ids,
        )
        .select_related("student")
        .order_by("student__last_name", "student__first_name")
    )


def _get_all_exchange_rows(student_ids: set, alt_map: dict) -> list[tuple]:
    """(gender, is_alternant, duration_weeks) pour TOUT l'historique d'échanges de la cohorte."""
    from app.outgoing.models import AssignmentResult, AssignmentStatus, SlotType

    qs = (
        AssignmentResult.objects.filter(
            assignment__status__in=[
                AssignmentStatus.VALIDATED,
                AssignmentStatus.PUBLISHED,
            ],
            annual_enrollment__student_id__in=student_ids,
        )
        .exclude(slot_type=SlotType.UNASSIGNED)
        .filter(agreement_year__isnull=False)
        .select_related("annual_enrollment__student", "agreement_year")
    )
    return [
        (
            r.annual_enrollment.student.gender or "M",
            alt_map.get(r.annual_enrollment.student_id, False),
            r.agreement_year.duration_weeks,
        )
        for r in qs
    ]


def _get_all_internship_rows(student_ids: set, alt_map: dict) -> list[tuple]:
    """(gender, is_alternant, weeks_in_company) pour TOUS les stages intl de la cohorte."""
    from app.internships.models import Internship

    qs = (
        Internship.objects.filter(student_id__in=student_ids)
        .exclude(country__isnull=True)
        .exclude(country__iso2="FR")
        .select_related("student")
    )
    return [
        (
            i.student.gender or "M",
            alt_map.get(i.student_id, False),
            i.weeks_in_company,
        )
        for i in qs
    ]


def _get_incoming_rows(academic_year: AcademicYear) -> list[tuple]:
    """(gender_cat, duration_cat) par étudiant entrant de l'année."""
    from app.incoming.models import IncomingStudent

    return [
        (_incoming_gender(s.civility), _parse_incoming_duration(s.duration))
        for s in IncomingStudent.objects.filter(academic_year=academic_year)
    ]


# ─── Statistiques VII-D3-3 (cohorte) ─────────────────────────────────────────


def _build_cohort_stats_3(student_ids: set, alt_map: dict, enrolled: list) -> dict:
    """% étudiants avec mobilité (3.a) et durée moyenne (3.b) sur tout l'historique."""
    from app.internships.models import Internship
    from app.outgoing.models import AssignmentResult, AssignmentStatus, SlotType

    base_exc = (
        AssignmentResult.objects.filter(
            assignment__status__in=[
                AssignmentStatus.VALIDATED,
                AssignmentStatus.PUBLISHED,
            ],
            annual_enrollment__student_id__in=student_ids,
        )
        .exclude(slot_type=SlotType.UNASSIGNED)
        .filter(agreement_year__isnull=False)
    )
    base_int = (
        Internship.objects.filter(student_id__in=student_ids)
        .exclude(country__isnull=True)
        .exclude(country__iso2="FR")
    )

    mobility_sids = set(
        base_exc.values_list("annual_enrollment__student_id", flat=True)
    ) | set(base_int.values_list("student_id", flat=True))

    fise_total = sum(1 for e in enrolled if not e.is_alternant)
    fisa_total = sum(1 for e in enrolled if e.is_alternant)
    fise_with = sum(1 for s in mobility_sids if not alt_map.get(s, False))
    fisa_with = sum(1 for s in mobility_sids if alt_map.get(s, False))

    exc_weeks: dict[int, float] = {}
    for r in base_exc.filter(
        agreement_year__duration_weeks__isnull=False
    ).select_related("annual_enrollment", "agreement_year"):
        sid = r.annual_enrollment.student_id
        exc_weeks[sid] = exc_weeks.get(sid, 0) + r.agreement_year.duration_weeks

    int_weeks: dict[int, float] = {}
    for i in base_int.filter(weeks_in_company__isnull=False):
        int_weeks[i.student_id] = int_weeks.get(i.student_id, 0) + i.weeks_in_company

    fise_w = [
        exc_weeks.get(s, 0) + int_weeks.get(s, 0)
        for s in mobility_sids
        if not alt_map.get(s, False)
    ]
    fisa_w = [
        exc_weeks.get(s, 0) + int_weeks.get(s, 0)
        for s in mobility_sids
        if alt_map.get(s, False)
    ]

    return {
        "fise_pct": round(fise_with / fise_total * 100, 1) if fise_total else 0.0,
        "fisa_pct": round(fisa_with / fisa_total * 100, 1) if fisa_total else 0.0,
        "fise_avg_months": _weeks_to_months(sum(fise_w) / len(fise_w))
        if fise_w
        else 0.0,
        "fisa_avg_months": _weeks_to_months(sum(fisa_w) / len(fisa_w))
        if fisa_w
        else 0.0,
    }


# ─── Construction des tableaux VII.D3 ─────────────────────────────────────────


def _count(rows, cat_fn):
    counts: dict = {}
    for gender, is_alt, value in rows:
        key = (gender, is_alt, cat_fn(value))
        counts[key] = counts.get(key, 0) + 1
    return counts


def _build_1a(exchange_rows):
    counts = _count(exchange_rows, _exchange_cat)

    def g(gender, is_alt, cat):
        return counts.get((gender, is_alt, cat), 0)

    col_headers = [
        "< 1 sem.\nFISE",
        "< 1 sem.\nFISA+FISEA",
        "1 semestre\nFISE",
        "1 semestre\nFISA+FISEA",
        "> 1 sem.\nFISE",
        "> 1 sem.\nFISA+FISEA",
    ]
    rows = [
        (
            "Hommes",
            [
                g("M", False, "lt"),
                g("M", True, "lt"),
                g("M", False, "eq"),
                g("M", True, "eq"),
                g("M", False, "gt"),
                g("M", True, "gt"),
            ],
        ),
        (
            "Femmes",
            [
                g("F", False, "lt"),
                g("F", True, "lt"),
                g("F", False, "eq"),
                g("F", True, "eq"),
                g("F", False, "gt"),
                g("F", True, "gt"),
            ],
        ),
        (
            "Total",
            [
                g("M", a, c) + g("F", a, c)
                for c in ("lt", "eq", "gt")
                for a in (False, True)
            ],
        ),
    ]
    return col_headers, rows


def _build_1b(internship_rows):
    counts = _count(internship_rows, _internship_cat)

    def g(gender, is_alt, cat):
        return counts.get((gender, is_alt, cat), 0)

    col_headers = [
        "< 3 mois\nFISE",
        "< 3 mois\nFISA+FISEA",
        "3–6 mois\nFISE",
        "3–6 mois\nFISA+FISEA",
        "> 6 mois\nFISE",
        "> 6 mois\nFISA+FISEA",
    ]
    rows = [
        (
            "Hommes",
            [
                g("M", False, "lt"),
                g("M", True, "lt"),
                g("M", False, "mid"),
                g("M", True, "mid"),
                g("M", False, "gt"),
                g("M", True, "gt"),
            ],
        ),
        (
            "Femmes",
            [
                g("F", False, "lt"),
                g("F", True, "lt"),
                g("F", False, "mid"),
                g("F", True, "mid"),
                g("F", False, "gt"),
                g("F", True, "gt"),
            ],
        ),
        (
            "Total",
            [
                g("M", a, c) + g("F", a, c)
                for c in ("lt", "mid", "gt")
                for a in (False, True)
            ],
        ),
    ]
    return col_headers, rows


def _build_4(incoming_rows):
    counts: dict = {}
    for gender, cat in incoming_rows:
        key = (gender, cat)
        counts[key] = counts.get(key, 0) + 1

    def g(gender, cat):
        return counts.get((gender, cat), 0)

    col_headers = ["< 1 semestre", "1 semestre", "> 1 semestre"]
    rows = [
        ("Hommes", [g("M", "lt"), g("M", "eq"), g("M", "gt")]),
        ("Femmes", [g("F", "lt"), g("F", "eq"), g("F", "gt")]),
        ("Total", [g("M", c) + g("F", c) for c in ("lt", "eq", "gt")]),
    ]
    return col_headers, rows


# ─── Écriture Excel ───────────────────────────────────────────────────────────


def _c(ws, row, col, value="", fill=None, font=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = align or _CENTER
    if border:
        cell.border = border
    return cell


def _title(ws, row, text, span=8):
    _c(ws, row, 1, text, fill=_TITLE_FILL, font=_TITLE_FONT, align=_LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18


def _section(ws, row, text, span=8):
    _c(ws, row, 1, text, fill=_SECTION_FILL, font=_SECTION_FONT, align=_LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 15


def _note(ws, row, text, span=8):
    _c(ws, row, 1, text, font=_NOTE_FONT, align=_LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _write_2row_table(ws, start_row, cat_labels, data_rows):
    """
    2 lignes d'en-têtes : catégories fusionnées (FISE + FISA+FISEA) sur la 1re,
    FISE / FISA+FISEA sur la 2e. Correspond au format CTI pour 1.a et 1.b.
    Colonnes : A=indent B=Durée cumulée/label C-D=cat1 E-F=cat2 G-H=cat3
    """
    r1, r2 = start_row, start_row + 1

    _c(ws, r1, 1, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(
        ws,
        r1,
        2,
        "Durée cumulée",
        fill=_COL_HDR_FILL,
        font=_COL_HDR_FONT,
        align=_LEFT,
        border=_BORDER,
    )
    for i, cat in enumerate(cat_labels):
        col = 3 + i * 2
        _c(ws, r1, col, cat, fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
        _c(ws, r1, col + 1, "", fill=_COL_HDR_FILL, border=_BORDER)
        ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + 1)
    ws.row_dimensions[r1].height = 22

    _c(ws, r2, 1, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(ws, r2, 2, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    for i in range(3):
        _c(
            ws,
            r2,
            3 + i * 2,
            "FISE",
            fill=_COL_HDR_FILL,
            font=_COL_HDR_FONT,
            border=_BORDER,
        )
        _c(
            ws,
            r2,
            3 + i * 2 + 1,
            "FISA + FISEA",
            fill=_COL_HDR_FILL,
            font=_COL_HDR_FONT,
            border=_BORDER,
        )
    ws.row_dimensions[r2].height = 15

    row = start_row + 2
    for label, values in data_rows:
        is_total = label == "Total"
        f, fn = (_TOTAL_FILL, _TOTAL_FONT) if is_total else (None, _DATA_FONT)
        _c(ws, row, 1, "", fill=f, font=fn, border=_BORDER)
        _c(ws, row, 2, label, fill=f, font=fn, align=_LEFT, border=_BORDER)
        for j, v in enumerate(values, 3):
            _c(ws, row, j, v, fill=f, font=fn, border=_BORDER)
        row += 1
    return row + 1


def _write_1row_table(ws, start_row, col_labels, data_rows):
    """1 ligne d'en-tête avec 'Durée' en col B. Utilisé pour VII-D3-4."""
    _c(ws, start_row, 1, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(
        ws,
        start_row,
        2,
        "Durée",
        fill=_COL_HDR_FILL,
        font=_COL_HDR_FONT,
        align=_LEFT,
        border=_BORDER,
    )
    for i, h in enumerate(col_labels, 3):
        _c(ws, start_row, i, h, fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    ws.row_dimensions[start_row].height = 22

    row = start_row + 1
    for label, values in data_rows:
        is_total = label == "Total"
        f, fn = (_TOTAL_FILL, _TOTAL_FONT) if is_total else (None, _DATA_FONT)
        _c(ws, row, 1, "", fill=f, font=fn, border=_BORDER)
        _c(ws, row, 2, label, fill=f, font=fn, align=_LEFT, border=_BORDER)
        for j, v in enumerate(values, 3):
            _c(ws, row, j, v, fill=f, font=fn, border=_BORDER)
        row += 1
    return row + 1


def _get_dd_incoming_rows(academic_year: AcademicYear) -> list[tuple]:
    """(region_label, gender) pour les étudiants entrants en DD (cadre DD)."""
    from app.incoming.models import IncomingStudent

    rows = []
    for s in IncomingStudent.objects.filter(
        academic_year=academic_year,
        mobility_category__name__icontains="dd",
    ).select_related("country", "mobility_category"):
        if not s.country or not s.country.cti_region:
            continue
        label = _CTI_REGION_LABELS.get(s.country.cti_region)
        if label:
            rows.append((label, _incoming_gender(s.civility)))
    return rows


def _get_dd_outgoing_rows(student_ids: set) -> list[tuple]:
    """(region_label, gender) pour les étudiants FISE sortants en DD — tout l'historique de la cohorte."""
    from app.outgoing.models import AssignmentResult, AssignmentStatus, SlotType

    rows = []
    for r in (
        AssignmentResult.objects.filter(
            assignment__status__in=[
                AssignmentStatus.VALIDATED,
                AssignmentStatus.PUBLISHED,
            ],
            annual_enrollment__student_id__in=student_ids,
            annual_enrollment__is_alternant=False,
            agreement_year__agreement__category__name__icontains="dd",
        )
        .exclude(slot_type=SlotType.UNASSIGNED)
        .select_related(
            "annual_enrollment__student",
            "agreement_year__agreement__partner_university__country",
        )
    ):
        ay = r.agreement_year
        if not ay or not ay.agreement or not ay.agreement.partner_university:
            continue
        country = ay.agreement.partner_university.country
        if not country or not country.cti_region:
            continue
        label = _CTI_REGION_LABELS.get(country.cti_region)
        if label:
            rows.append((label, r.annual_enrollment.student.gender or "M"))
    return rows


def _write_region_table(ws, start_row, dd_rows: list | None = None):
    counts: dict = {}
    for region_label, gender in dd_rows or []:
        counts[(region_label, gender)] = counts.get((region_label, gender), 0) + 1

    for i, h in enumerate(
        ["Pays d'obtention de l'autre diplôme", "Hommes", "Femmes", "Total"], 1
    ):
        _c(
            ws,
            start_row,
            i,
            h,
            fill=_COL_HDR_FILL,
            font=_COL_HDR_FONT,
            align=_LEFT if i == 1 else _CENTER,
            border=_BORDER,
        )
    row = start_row + 1
    for region in _REGIONS:
        h_count = counts.get((region, "M"), 0)
        f_count = counts.get((region, "F"), 0)
        _c(ws, row, 1, region, font=_DATA_FONT, align=_LEFT, border=_BORDER)
        _c(ws, row, 2, h_count, font=_DATA_FONT, border=_BORDER)
        _c(ws, row, 3, f_count, font=_DATA_FONT, border=_BORDER)
        _c(ws, row, 4, h_count + f_count, font=_DATA_FONT, border=_BORDER)
        row += 1
    return row + 1


_REGIONS = [
    "Europe (hors France)",
    "Canada / États-Unis",
    "Autres pays d'Amérique",
    "Pays d'Asie y compris Moyen Orient",
    "Pays d'Afrique",
    "Océanie",
]
_CTI_REGION_LABELS = {
    "europe_hors_france": "Europe (hors France)",
    "canada_usa": "Canada / États-Unis",
    "amerique": "Autres pays d'Amérique",
    "asie_moyen_orient": "Pays d'Asie y compris Moyen Orient",
    "afrique": "Pays d'Afrique",
    "oceanie": "Océanie",
}


# ─── Feuille VII.D3 ──────────────────────────────────────────────────────────


def _sheet_vii_d3(
    ws,
    title: str,
    exchange_rows,
    internship_rows,
    incoming_rows,
    stats,
    dd_incoming_rows=None,
    dd_outgoing_rows=None,
):
    span = 9
    row = 1

    _title(ws, row, f"RAPPORT CTI — VII.D3 — {title}", span=span)
    row += 2

    # ── MOBILITÉ SORTANTE ─────────────────────────────────────────────────────
    _section(ws, row, "MOBILITÉ SORTANTE", span=span)
    row += 1
    _note(
        ws,
        row,
        "Nombre de diplômés de la dernière promotion ayant vécu une expérience à l'étranger dans le cadre de leur formation",
        span=span,
    )
    row += 2

    # 1.a
    _section(
        ws,
        row,
        "VII-D3-1.a — Diplômés de la dernière promotion ayant effectué une ou plusieurs mobilités académiques au cours de leur scolarité",
        span=span,
    )
    row += 1
    _, exc_rows = _build_1a(exchange_rows)
    row = _write_2row_table(
        ws,
        row,
        [
            "Moins d'un semestre",
            "1 semestre",
            "Plus d'un semestre (en continu ou non)",
        ],
        exc_rows,
    )

    # 1.b
    _section(
        ws,
        row,
        "VII-D3-1.b — Diplômés de la dernière promotion ayant effectué un ou plusieurs stages à l'étranger",
        span=span,
    )
    row += 1
    _, int_rows = _build_1b(internship_rows)
    row = _write_2row_table(
        ws,
        row,
        [
            "< à 3 mois",
            "≥ à 3 mois et < à 6 mois",
            "> à 6 mois",
        ],
        int_rows,
    )

    # 2
    _section(
        ws,
        row,
        "VII-D3-2 — Doubles diplômés ingénieurs sortants (FISE seulement)",
        span=span,
    )
    row += 1
    row = _write_region_table(ws, row, dd_outgoing_rows)
    row += 1

    # Synthèse
    _section(ws, row, "Synthèse de la mobilité sortante FISE seulement", span=span)
    row += 2

    # 3.a
    _section(
        ws,
        row,
        "VII-D3-3.a — Pourcentage de diplômés ayant effectué une mobilité sortante à l'étranger (d'études ou de stage) au cours de leur formation",
        span=span,
    )
    row += 1
    _c(ws, row, 1, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(ws, row, 2, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(ws, row, 3, f"{stats['fise_pct']:.0f} %", font=_DATA_FONT, border=_BORDER)
    _c(
        ws,
        row,
        4,
        "On prend ici en compte tous les diplômés de l'école comptés dans les questions VII.1.a, VII.1.b et en VII.2",
        font=_NOTE_FONT,
        align=_LEFT,
    )
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=span)
    row += 2

    # 3.b
    _section(
        ws,
        row,
        "VII-D3-3.b — Durée moyenne de la mobilité à l'étranger parmi les diplômés comptabilisés en VII.3.a",
        span=span,
    )
    row += 1
    _c(ws, row, 1, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(ws, row, 2, "", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(ws, row, 3, "FISE", fill=_COL_HDR_FILL, font=_COL_HDR_FONT, border=_BORDER)
    _c(
        ws,
        row,
        4,
        "FISA + FISEA",
        fill=_COL_HDR_FILL,
        font=_COL_HDR_FONT,
        border=_BORDER,
    )
    _c(ws, row, 5, "La réponse est à donner en mois.", font=_NOTE_FONT, align=_LEFT)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=span)
    row += 1
    _c(ws, row, 1, "", border=_BORDER)
    _c(ws, row, 2, "", border=_BORDER)
    _c(ws, row, 3, f"{stats['fise_avg_months']} MOIS", font=_DATA_FONT, border=_BORDER)
    _c(ws, row, 4, f"{stats['fisa_avg_months']} MOIS", font=_DATA_FONT, border=_BORDER)
    row += 3

    # ── MOBILITÉ ENTRANTE ─────────────────────────────────────────────────────
    _section(ws, row, "MOBILITÉ ENTRANTE FISE SEULEMENT", span=span)
    row += 1
    _note(
        ws,
        row,
        f"Élèves étrangers en échange académique en provenance de l'étranger — {title}",
        span=span,
    )
    row += 2

    # 4
    _section(ws, row, "VII-D3-4", span=span)
    row += 1
    _, inc_rows = _build_4(incoming_rows)
    row = _write_1row_table(
        ws,
        row,
        [
            "Moins d'un semestre",
            "1 semestre",
            "Plus d'un semestre (en continu ou non)",
        ],
        inc_rows,
    )

    # 5
    _section(
        ws,
        row,
        "VII-D3-5 — Doubles diplômés ingénieurs entrants de la dernière promotion",
        span=span,
    )
    row += 1
    row = _write_region_table(ws, row, dd_incoming_rows)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    for letter in "CDEFGH":
        ws.column_dimensions[letter].width = 20
    ws.column_dimensions["I"].width = 40


# ─── Point d'entrée stats JSON ───────────────────────────────────────────────


def _region_counts(dd_rows: list) -> list[dict]:
    counts: dict = {}
    for region_label, gender in dd_rows or []:
        counts[(region_label, gender)] = counts.get((region_label, gender), 0) + 1
    return [
        {
            "region": region,
            "hommes": counts.get((region, "M"), 0),
            "femmes": counts.get((region, "F"), 0),
            "total": counts.get((region, "M"), 0) + counts.get((region, "F"), 0),
        }
        for region in _REGIONS
    ]


def get_cti_stats_data(academic_year: AcademicYear) -> dict:
    """
    Statistiques VII.D3 pour la cohorte (niveaux is_terminal=True) de l'année N.
    Mobilité collectée sur TOUT l'historique de chaque étudiant de la cohorte.
    """
    enrolled = _get_terminal_enrolled(academic_year)
    student_ids = {e.student_id for e in enrolled}
    alt_map = {e.student_id: e.is_alternant for e in enrolled}

    exchange_rows = _get_all_exchange_rows(student_ids, alt_map)
    internship_rows = _get_all_internship_rows(student_ids, alt_map)
    incoming_rows = _get_incoming_rows(academic_year)
    stats = _build_cohort_stats_3(student_ids, alt_map, enrolled)
    dd_outgoing_rows = _get_dd_outgoing_rows(student_ids)
    dd_incoming_rows = _get_dd_incoming_rows(academic_year)

    _, exc_data = _build_1a(exchange_rows)
    _, int_data = _build_1b(internship_rows)
    _, inc_data = _build_4(incoming_rows)

    def _to_exc(label, vals):
        return {
            "label": label,
            "lt_fise": vals[0],
            "lt_fisa": vals[1],
            "eq_fise": vals[2],
            "eq_fisa": vals[3],
            "gt_fise": vals[4],
            "gt_fisa": vals[5],
        }

    def _to_int(label, vals):
        return {
            "label": label,
            "lt_fise": vals[0],
            "lt_fisa": vals[1],
            "mid_fise": vals[2],
            "mid_fisa": vals[3],
            "gt_fise": vals[4],
            "gt_fisa": vals[5],
        }

    def _to_inc(label, vals):
        return {"label": label, "lt": vals[0], "eq": vals[1], "gt": vals[2]}

    return {
        "academic_year_label": academic_year.label,
        "enrolled_fise": sum(1 for e in enrolled if not e.is_alternant),
        "enrolled_fisa": sum(1 for e in enrolled if e.is_alternant),
        "exchanges": [_to_exc(lbl, v) for lbl, v in exc_data],
        "internships": [_to_int(lbl, v) for lbl, v in int_data],
        "pct_fise": stats["fise_pct"],
        "pct_fisa": stats["fisa_pct"],
        "avg_months_fise": stats["fise_avg_months"],
        "avg_months_fisa": stats["fisa_avg_months"],
        "incoming": [_to_inc(lbl, v) for lbl, v in inc_data],
        "dd_outgoing": _region_counts(dd_outgoing_rows),
        "dd_incoming": _region_counts(dd_incoming_rows),
    }


# ─── Point d'entrée Excel ────────────────────────────────────────────────────


def generate_cti_excel(academic_year: AcademicYear) -> bytes:
    """
    Génère le rapport CTI Excel (format VII.D3) pour la cohorte de l'année N.
    Retourne les bytes du fichier .xlsx.
    """
    enrolled = _get_terminal_enrolled(academic_year)
    student_ids = {e.student_id for e in enrolled}
    alt_map = {e.student_id: e.is_alternant for e in enrolled}

    exchange_rows = _get_all_exchange_rows(student_ids, alt_map)
    internship_rows = _get_all_internship_rows(student_ids, alt_map)
    incoming_rows = _get_incoming_rows(academic_year)
    stats = _build_cohort_stats_3(student_ids, alt_map, enrolled)

    dd_incoming_rows = _get_dd_incoming_rows(academic_year)
    dd_outgoing_rows = _get_dd_outgoing_rows(student_ids)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "VII.D3 CTI"
    _sheet_vii_d3(
        ws1,
        academic_year.label,
        exchange_rows,
        internship_rows,
        incoming_rows,
        stats,
        dd_incoming_rows,
        dd_outgoing_rows,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
