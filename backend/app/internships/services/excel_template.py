from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

from app.academic.models import AcademicYear
from app.shared.excel_utils import write_header_row

_HEADERS = [
    "Étudiant (INE – Nom Prénom)",
    "Raison sociale",
    "Pays",
    "Ville",
    "Date de début",
    "Date de fin",
    "Nb semaines",
    "Type",
    "Titre",
    "Tuteur pédagogique",
    "Tuteur entreprise",
]

_WIDTHS = [45, 35, 25, 20, 14, 14, 12, 16, 40, 30, 30]


def generate_internship_template(academic_year: AcademicYear) -> bytes:
    from app.reference.models import Country
    from app.students.models import AnnualEnrollment

    # Étudiants inscrits pour cette année
    enrollments = (
        AnnualEnrollment.objects.filter(
            academic_year=academic_year, deleted_at__isnull=True
        )
        .select_related("student")
        .order_by("student__last_name", "student__first_name")
    )
    student_rows = [
        (e.student.ine, e.student.last_name, e.student.first_name) for e in enrollments
    ]

    # Pays triés par nom français
    countries = list(
        Country.objects.order_by("name_fr").values_list("name_fr", flat=True)
    )

    wb = openpyxl.Workbook()

    # ── Feuille cachée : liste des étudiants ──
    ref_stu_ws = wb.create_sheet("_Etudiants")
    for ine, nom, prenom in student_rows:
        ref_stu_ws.append([f"{ine} – {nom} {prenom}"])
    ref_stu_ws.sheet_state = "hidden"

    # ── Feuille cachée : liste des pays ──
    ref_pays_ws = wb.create_sheet("_Pays")
    for name in countries:
        ref_pays_ws.append([name])
    ref_pays_ws.sheet_state = "hidden"

    # ── Feuille principale ──
    ws = wb.active
    ws.title = "Stages"

    write_header_row(ws, _HEADERS, _WIDTHS)

    # Validation colonne A → étudiants
    if student_rows:
        dv_stu = DataValidation(
            type="list",
            formula1=f"_Etudiants!$A$1:$A${len(student_rows)}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_stu.sqref = "A2:A10000"
        ws.add_data_validation(dv_stu)

    # Validation colonne C → pays
    if countries:
        dv_pays = DataValidation(
            type="list",
            formula1=f"_Pays!$A$1:$A${len(countries)}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_pays.sqref = "C2:C10000"
        ws.add_data_validation(dv_pays)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
