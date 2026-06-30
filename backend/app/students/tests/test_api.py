from datetime import date
from io import BytesIO

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client

from app.academic.models import AcademicYear
from app.imports.models import RawImport, RawImportEntity, RawImportStatus
from app.institutions.models import PartnerUniversity
from app.mobility.models import Agreement, AgreementYear
from app.outgoing.models import Assignment, AssignmentResult, AssignmentStatus, SlotType
from app.reference.models import Country, CTIRegion, Department, Level
from app.students.models import AnnualEnrollment, Student

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["INE", "Nom", "Prenom", "Email", "Genre", "Departement", "Niveau", "GPA"]
    )
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_file(rows: list[list], name: str = "etudiants.xlsx") -> SimpleUploadedFile:
    return SimpleUploadedFile(
        name,
        make_xlsx(rows),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def make_year(**kwargs) -> AcademicYear:
    defaults = {
        "label": "2026-2027",
        "start_date": date(2026, 9, 1),
        "end_date": date(2027, 8, 31),
    }
    defaults.update(kwargs)
    return AcademicYear.objects.create(**defaults)


# ---------------------------------------------------------------------------
# GET /students/students/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestListStudents:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        self.dept = Department.objects.create(code="SN", name="Sciences du Numerique")
        self.level = Level.objects.create(code="3A", name="Troisieme annee")

    def test_list_students_empty(self):
        response = self.client.get("/api/v1/students/students/")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 0
        assert body["results"] == []

    def test_list_students(self):
        student = Student.objects.create(
            ine="12345678901", first_name="Jean", last_name="Martin"
        )
        AnnualEnrollment.objects.create(
            student=student,
            academic_year=self.year,
            department=self.dept,
            level=self.level,
        )

        response = self.client.get("/api/v1/students/students/")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "12345678901"

    def test_filter_by_academic_year(self):
        year2 = make_year(
            label="2027-2028", start_date=date(2027, 9, 1), end_date=date(2028, 8, 31)
        )
        s1 = Student.objects.create(ine="10000000001", first_name="A", last_name="A")
        s2 = Student.objects.create(ine="10000000002", first_name="B", last_name="B")
        AnnualEnrollment.objects.create(
            student=s1, academic_year=self.year, department=self.dept, level=self.level
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=year2, department=self.dept, level=self.level
        )

        response = self.client.get(
            f"/api/v1/students/students/?academic_year_id={self.year.id}"
        )

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"

    def test_filter_by_department(self):
        dept2 = Department.objects.create(code="TC", name="Tronc Commun")
        s1 = Student.objects.create(ine="10000000001", first_name="A", last_name="A")
        s2 = Student.objects.create(ine="10000000002", first_name="B", last_name="B")
        AnnualEnrollment.objects.create(
            student=s1, academic_year=self.year, department=self.dept, level=self.level
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=self.year, department=dept2, level=self.level
        )

        response = self.client.get(
            f"/api/v1/students/students/?department_id={self.dept.id}"
        )

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"

    def test_filter_by_level(self):
        level2 = Level.objects.create(code="4A", name="Quatrieme annee")
        s1 = Student.objects.create(ine="10000000001", first_name="A", last_name="A")
        s2 = Student.objects.create(ine="10000000002", first_name="B", last_name="B")
        AnnualEnrollment.objects.create(
            student=s1, academic_year=self.year, department=self.dept, level=self.level
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=self.year, department=self.dept, level=level2
        )

        response = self.client.get(
            f"/api/v1/students/students/?level_id={self.level.id}"
        )

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"

    def test_search_students(self):
        Student.objects.create(ine="10000000001", first_name="Jean", last_name="Martin")
        Student.objects.create(
            ine="10000000002", first_name="Marie", last_name="Dupont"
        )

        response = self.client.get("/api/v1/students/students/?search=martin")

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"


# ---------------------------------------------------------------------------
# GET /students/students/select-options/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestStudentSelectOptions:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        self.dept = Department.objects.create(code="SN", name="Sciences du Numerique")
        self.level = Level.objects.create(code="3A", name="Troisieme annee")

    def test_select_options_empty(self):
        response = self.client.get("/api/v1/students/students/select-options/")

        assert response.status_code == 200
        assert response.json() == []

    def test_select_options_with_year_filter(self):
        student = Student.objects.create(
            ine="12345678901", first_name="Jean", last_name="Martin"
        )
        AnnualEnrollment.objects.create(
            student=student,
            academic_year=self.year,
            department=self.dept,
            level=self.level,
        )

        response = self.client.get(
            f"/api/v1/students/students/select-options/?academic_year_id={self.year.id}"
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert "Martin" in data[0]["label"]

    def test_select_options_with_search(self):
        Student.objects.create(ine="10000000001", first_name="Jean", last_name="Martin")
        Student.objects.create(
            ine="10000000002", first_name="Marie", last_name="Dupont"
        )

        response = self.client.get(
            "/api/v1/students/students/select-options/?search=dupont"
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert "Dupont" in data[0]["label"]


# ---------------------------------------------------------------------------
# GET /students/students/{id}/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestGetStudent:
    def setup_method(self):
        self.client = Client()
        self.student = Student.objects.create(
            ine="12345678901", first_name="Jean", last_name="Martin"
        )

    def test_get_student(self):
        response = self.client.get(f"/api/v1/students/students/{self.student.id}/")

        assert response.status_code == 200
        data = response.json()
        assert data["ine"] == "12345678901"
        assert "enrollments" in data

    def test_get_student_not_found(self):
        response = self.client.get("/api/v1/students/students/99999/")

        assert response.status_code == 404


# ---------------------------------------------------------------------------
# GET /students/students/by-year/{year_id}/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestListStudentsByYear:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        self.dept = Department.objects.create(code="SN", name="Sciences du Numerique")
        self.level = Level.objects.create(code="3A", name="Troisieme annee")

    def test_by_year_empty(self):
        response = self.client.get(f"/api/v1/students/students/by-year/{self.year.id}/")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 0
        assert body["results"] == []

    def test_by_year_returns_enrollment_details(self):
        student = Student.objects.create(
            ine="12345678901", first_name="Jean", last_name="Martin", gender="M"
        )
        AnnualEnrollment.objects.create(
            student=student,
            academic_year=self.year,
            department=self.dept,
            level=self.level,
            gpa=15.5,
        )

        response = self.client.get(f"/api/v1/students/students/by-year/{self.year.id}/")

        assert response.status_code == 200
        body = response.json()
        assert body["count"] == 1
        row = body["results"][0]
        assert row["ine"] == "12345678901"
        assert row["department_code"] == "SN"
        assert row["level_code"] == "3A"
        assert row["gpa"] is not None
        assert row["parcours_code"] is None

    def test_by_year_ordered_by_last_name(self):
        for ine, last_name in [("10000000001", "Zeta"), ("10000000002", "Alpha")]:
            s = Student.objects.create(ine=ine, first_name="X", last_name=last_name)
            AnnualEnrollment.objects.create(
                student=s,
                academic_year=self.year,
                department=self.dept,
                level=self.level,
            )

        response = self.client.get(f"/api/v1/students/students/by-year/{self.year.id}/")

        results = response.json()["results"]
        assert results[0]["last_name"] == "Alpha"
        assert results[1]["last_name"] == "Zeta"

    def test_filter_by_level(self):
        level2 = Level.objects.create(code="4A", name="Quatrieme annee")
        s1 = Student.objects.create(ine="10000000001", first_name="A", last_name="A")
        s2 = Student.objects.create(ine="10000000002", first_name="B", last_name="B")
        AnnualEnrollment.objects.create(
            student=s1, academic_year=self.year, department=self.dept, level=self.level
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=self.year, department=self.dept, level=level2
        )

        response = self.client.get(
            f"/api/v1/students/students/by-year/{self.year.id}/?level_id={self.level.id}"
        )

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"

    def test_filter_by_parcours(self):
        from app.reference.models import Parcours

        parcours = Parcours.objects.create(
            department=self.dept, code="SESG", label="Systemes Embarques"
        )
        s1 = Student.objects.create(ine="10000000001", first_name="A", last_name="A")
        s2 = Student.objects.create(ine="10000000002", first_name="B", last_name="B")
        AnnualEnrollment.objects.create(
            student=s1,
            academic_year=self.year,
            department=self.dept,
            level=self.level,
            parcours=parcours,
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=self.year, department=self.dept, level=self.level
        )

        response = self.client.get(
            f"/api/v1/students/students/by-year/{self.year.id}/?parcours_id={parcours.id}"
        )

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"

    def test_search_by_year(self):
        s1 = Student.objects.create(
            ine="10000000001", first_name="Jean", last_name="Martin"
        )
        s2 = Student.objects.create(
            ine="10000000002", first_name="Marie", last_name="Dupont"
        )
        AnnualEnrollment.objects.create(
            student=s1, academic_year=self.year, department=self.dept, level=self.level
        )
        AnnualEnrollment.objects.create(
            student=s2, academic_year=self.year, department=self.dept, level=self.level
        )

        response = self.client.get(
            f"/api/v1/students/students/by-year/{self.year.id}/?search=martin"
        )

        body = response.json()
        assert body["count"] == 1
        assert body["results"][0]["ine"] == "10000000001"


# ---------------------------------------------------------------------------
# GET /students/students/stats/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestStudentStats:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        self.dept = Department.objects.create(code="SN", name="Sciences du Numerique")
        self.level = Level.objects.create(code="3A", name="Troisieme annee")

    def test_stats_empty_year(self):
        response = self.client.get(
            f"/api/v1/students/students/stats/?academic_year_id={self.year.id}"
        )

        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 0
        assert data["by_level"] == []
        assert data["by_department"] == []
        assert data["by_parcours"] == []

    def test_stats_counts_enrollments(self):
        for i in range(3):
            s = Student.objects.create(
                ine=f"1000000000{i}", first_name="A", last_name="B"
            )
            AnnualEnrollment.objects.create(
                student=s,
                academic_year=self.year,
                department=self.dept,
                level=self.level,
            )

        response = self.client.get(
            f"/api/v1/students/students/stats/?academic_year_id={self.year.id}"
        )

        data = response.json()
        assert data["total"] == 3
        assert data["by_level"][0]["level_code"] == "3A"
        assert data["by_level"][0]["count"] == 3
        assert data["by_department"][0]["department_code"] == "SN"
        assert data["by_department"][0]["count"] == 3

    def test_stats_cross_breakdown_present(self):
        student = Student.objects.create(
            ine="12345678901", first_name="A", last_name="B"
        )
        AnnualEnrollment.objects.create(
            student=student,
            academic_year=self.year,
            department=self.dept,
            level=self.level,
        )

        response = self.client.get(
            f"/api/v1/students/students/stats/?academic_year_id={self.year.id}"
        )

        data = response.json()
        assert len(data["cross"]) == 1
        assert data["cross"][0]["level_code"] == "3A"
        assert data["cross"][0]["department_code"] == "SN"


# ---------------------------------------------------------------------------
# GET /students/students/import-errors/
# PUT /students/students/import-errors/{id}/ignore/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestStudentImportErrors:
    def setup_method(self):
        self.client = Client()

    def test_list_empty(self):
        response = self.client.get("/api/v1/students/students/import-errors/")

        assert response.status_code == 200
        assert response.json()["results"] == []

    def test_list_returns_only_failed(self):
        RawImport.objects.create(
            source="pegase",
            entity=RawImportEntity.STUDENT,
            external_id="12345678901",
            payload={"ine": "12345678901"},
            status=RawImportStatus.FAILED,
            error_message="Département introuvable: INCONNU",
        )
        RawImport.objects.create(
            source="pegase",
            entity=RawImportEntity.STUDENT,
            external_id="12345678902",
            payload={"ine": "12345678902"},
            status=RawImportStatus.IMPORTED,
        )

        response = self.client.get("/api/v1/students/students/import-errors/")

        data = response.json()
        assert data["count"] == 1
        assert data["results"][0]["external_id"] == "12345678901"

    def test_list_deduplicates_by_ine(self):
        for _ in range(3):
            RawImport.objects.create(
                source="pegase",
                entity=RawImportEntity.STUDENT,
                external_id="12345678901",
                payload={"ine": "12345678901"},
                status=RawImportStatus.FAILED,
                error_message="Niveau introuvable: X",
            )

        response = self.client.get("/api/v1/students/students/import-errors/")

        assert response.json()["count"] == 1

    def test_list_excludes_other_entities(self):
        RawImport.objects.create(
            source="pegase",
            entity=RawImportEntity.DEPARTMENT,
            external_id="101",
            payload={"code": "XX"},
            status=RawImportStatus.FAILED,
            error_message="code manquant",
        )

        response = self.client.get("/api/v1/students/students/import-errors/")

        assert response.json()["results"] == []

    def test_ignore_error(self):
        raw = RawImport.objects.create(
            source="pegase",
            entity=RawImportEntity.STUDENT,
            external_id="12345678901",
            payload={"ine": "12345678901"},
            status=RawImportStatus.FAILED,
            error_message="Département introuvable: INCONNU",
        )

        response = self.client.put(
            f"/api/v1/students/students/import-errors/{raw.id}/ignore/"
        )

        assert response.status_code == 200
        assert response.json()["status"] == RawImportStatus.IGNORED
        raw.refresh_from_db()
        assert raw.status == RawImportStatus.IGNORED

    def test_ignore_nonexistent_returns_404(self):
        response = self.client.put(
            "/api/v1/students/students/import-errors/99999/ignore/"
        )

        assert response.status_code == 404

    def test_ignore_already_imported_returns_404(self):
        raw = RawImport.objects.create(
            source="pegase",
            entity=RawImportEntity.STUDENT,
            external_id="12345678901",
            payload={"ine": "12345678901"},
            status=RawImportStatus.IMPORTED,
        )

        response = self.client.put(
            f"/api/v1/students/students/import-errors/{raw.id}/ignore/"
        )

        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /students/students/sync-pegase/{year_id}/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestSyncPegase:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        Department.objects.create(code="SN", name="Sciences du Numerique")
        Level.objects.create(code="3A", name="Troisieme annee")

    def test_sync_triggers_background_task(self, monkeypatch):
        import app.students.api as students_api

        calls = []
        monkeypatch.setattr(
            students_api,
            "enqueue_sync_pegase_students",
            lambda year_id, triggered_by="": calls.append(year_id) or "fake-task-id",
        )

        response = self.client.post(
            f"/api/v1/students/students/sync-pegase/{self.year.id}/"
        )

        assert response.status_code == 202
        data = response.json()
        assert data["task_id"] == "fake-task-id"
        assert calls == [self.year.id]

    def test_sync_unknown_year_returns_404(self):
        response = self.client.post("/api/v1/students/students/sync-pegase/99999/")

        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /students/students/import-excel/{year_id}/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestImportExcel:
    def setup_method(self):
        self.client = Client()
        self.year = make_year()
        Department.objects.create(code="SN", name="Sciences du Numerique")
        Level.objects.create(code="3A", name="Troisieme annee")

    def test_import_triggers_background_task(self, monkeypatch):
        import app.students.api as students_api

        calls = []
        monkeypatch.setattr(
            students_api,
            "enqueue_import_excel_students",
            lambda file_bytes, source_file, year_id, triggered_by="": (
                calls.append(year_id) or "fake-task-id"
            ),
        )

        response = self.client.post(
            f"/api/v1/students/students/import-excel/{self.year.id}/",
            data={"file": xlsx_file([])},
        )

        assert response.status_code == 202
        data = response.json()
        assert data["task_id"] == "fake-task-id"
        assert calls == [self.year.id]

    def test_import_unknown_year_returns_404(self):
        response = self.client.post(
            "/api/v1/students/students/import-excel/99999/",
            data={"file": xlsx_file([])},
        )

        assert response.status_code == 404


# ---------------------------------------------------------------------------
# GET /students/students/template/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestTemplateDownload:
    def setup_method(self):
        self.client = Client()
        Department.objects.create(code="SN", name="Sciences du Numerique")
        Level.objects.create(code="3A", name="Troisieme annee")

    def test_download_returns_xlsx(self):
        response = self.client.get("/api/v1/students/students/template/")

        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "template_etudiants.xlsx" in response["Content-Disposition"]

    def test_download_is_valid_workbook(self):
        response = self.client.get("/api/v1/students/students/template/")

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "INE" in headers
        assert "GPA" in headers


# ---------------------------------------------------------------------------
# GET /api/v1/student/{ine}/assignment/
# ---------------------------------------------------------------------------


def _make_published_assignment(ine: str) -> AssignmentResult:
    year = AcademicYear.objects.create(
        label="2026-2027", start_date=date(2026, 9, 1), end_date=date(2027, 8, 31)
    )
    dept = Department.objects.create(code="SN", name="Sciences du Numerique")
    level = Level.objects.create(code="3A", name="Troisieme annee")
    student = Student.objects.create(ine=ine, first_name="Alice", last_name="Dupont")
    enrollment = AnnualEnrollment.objects.create(
        student=student, academic_year=year, department=dept, level=level
    )
    assignment = Assignment.objects.create(academic_year=year)
    Assignment.objects.filter(pk=assignment.pk).update(
        status=AssignmentStatus.PUBLISHED
    )
    assignment = Assignment.objects.get(pk=assignment.pk)
    return enrollment, assignment, year


def _make_agreement_year(year: AcademicYear) -> AgreementYear:
    country = Country.objects.create(
        iso2="IT",
        name_fr="Italie",
        name_en="Italy",
        cti_region=CTIRegion.EUROPE_HORS_FRANCE,
    )
    univ = PartnerUniversity.objects.create(
        name="Politecnico di Milano", short_name="PoliMi", country=country
    )
    agreement = Agreement.objects.create(
        name="Erasmus+ Polimi",
        partner_university=univ,
        direction="outgoing",
        inp_total_places=5,
    )
    return AgreementYear.objects.create(
        agreement=agreement, academic_year=year, is_active=True, n7_places=5
    )


@pytest.mark.django_db
class TestGetStudentAssignment:
    def setup_method(self):
        self.client = Client()

    def test_returns_404_when_no_published_assignment(self):
        response = self.client.get("/api/v1/student/UNKNOWN123/assignment/")
        assert response.status_code == 404

    def test_returns_assigned_result(self):
        enrollment, assignment, year = _make_published_assignment("20SN001")
        ay = _make_agreement_year(year)
        AssignmentResult.objects.create(
            assignment=assignment,
            annual_enrollment=enrollment,
            slot_type=SlotType.DEPT,
            agreement_year=ay,
            assigned_rank=1,
        )

        response = self.client.get("/api/v1/student/20SN001/assignment/")

        assert response.status_code == 200
        data = response.json()
        assert data["is_assigned"] is True
        assert data["slot_type"] == "dept"
        assert "Politecnico di Milano" in data["university_name"]
        assert data["country_name"] == "Italie"
        assert data["agreement_name"] == "Erasmus+ Polimi"
        assert data["effective_rank"] == 1

    def test_returns_unassigned_result(self):
        enrollment, assignment, _ = _make_published_assignment("20SN002")
        AssignmentResult.objects.create(
            assignment=assignment,
            annual_enrollment=enrollment,
            slot_type=SlotType.UNASSIGNED,
        )

        response = self.client.get("/api/v1/student/20SN002/assignment/")

        assert response.status_code == 200
        data = response.json()
        assert data["is_assigned"] is False
        assert data["university_name"] is None

    def test_override_agreement_takes_precedence(self):
        enrollment, assignment, year = _make_published_assignment("20SN003")
        ay_auto = _make_agreement_year(year)
        country2 = Country.objects.create(
            iso2="ES",
            name_fr="Espagne",
            name_en="Spain",
            cti_region=CTIRegion.EUROPE_HORS_FRANCE,
        )
        univ2 = PartnerUniversity.objects.create(
            name="UPM Madrid", short_name="UPM", country=country2
        )
        agreement2 = Agreement.objects.create(
            name="Erasmus+ UPM",
            partner_university=univ2,
            direction="outgoing",
            inp_total_places=3,
        )
        ay_override = AgreementYear.objects.create(
            agreement=agreement2, academic_year=year, is_active=True, n7_places=3
        )
        AssignmentResult.objects.create(
            assignment=assignment,
            annual_enrollment=enrollment,
            slot_type=SlotType.DEPT,
            agreement_year=ay_auto,
            override_agreement_year=ay_override,
            override_reason="Meilleure adéquation pédagogique",
            source="override",
        )

        response = self.client.get("/api/v1/student/20SN003/assignment/")

        assert response.status_code == 200
        data = response.json()
        assert data["is_assigned"] is True
        assert "UPM Madrid" in data["university_name"]
        assert data["country_name"] == "Espagne"
        assert data["override_reason"] == "Meilleure adéquation pédagogique"
