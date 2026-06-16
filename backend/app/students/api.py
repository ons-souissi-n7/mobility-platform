from io import BytesIO

import openpyxl
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from ninja import File, Query, Router
from ninja.errors import HttpError
from ninja.files import UploadedFile
from openpyxl.styles import Alignment, Font, PatternFill

from app.academic.api import get_academic_year
from app.audit.logger import log_action
from app.imports.models import (
    ImportReport as DbImportReport,
)
from app.imports.models import (
    ImportSource,
    RawImport,
    RawImportEntity,
    RawImportStatus,
)
from app.mobility.models import Agreement
from app.reference.models import Department, Level, Parcours
from app.shared.api_helpers import (
    PagedResponse,
    PaginationQuery,
    SelectOption,
    paginate,
)
from app.shared.excel_utils import build_filename, workbook_response, write_header_row
from app.shared.validators import DomainValidationError

from .models import AnnualEnrollment, Student, StudentWish
from .schemas import (
    AgreementWishOut,
    CrossStatOut,
    DepartmentStatOut,
    ImportReportOut,
    LevelStatOut,
    ParcoursStatOut,
    StudentDetailOut,
    StudentEnrollmentOut,
    StudentImportRetryIn,
    StudentOut,
    StudentRawImportOut,
    StudentStatsOut,
    StudentWishesOut,
    WishImportRetryIn,
    WishSyncReportOut,
)
from .services.adapters import excel as excel_adapter
from .services.adapters import excel_wishes as excel_wishes_adapter
from .services.adapters import pegase as pegase_adapter
from .services.student_importer import StudentRow, import_students
from .services.student_transformer import transform_student, transform_wish
from .services.student_validator import validate_student, validate_wish
from .services.sync_moveon import WishRow, import_wish_rows, sync_moveon_wishes

router = Router()


@router.get(
    "/students/", response=PagedResponse[StudentOut], summary="Liste des etudiants"
)
def list_students(
    request,
    academic_year_id: int | None = None,
    department_id: int | None = None,
    level_id: int | None = None,
    search: str | None = None,
    pagination: PaginationQuery = Query(),
):
    qs = Student.objects.all()
    if academic_year_id:
        qs = qs.filter(enrollments__academic_year_id=academic_year_id)
    if department_id:
        qs = qs.filter(enrollments__department_id=department_id)
    if level_id:
        qs = qs.filter(enrollments__level_id=level_id)
    if search:
        qs = qs.filter(
            Q(last_name__icontains=search)
            | Q(first_name__icontains=search)
            | Q(ine__icontains=search)
        )
    qs = qs.distinct()
    count, items = paginate(qs, pagination.page, pagination.page_size)
    return PagedResponse(
        count=count, page=pagination.page, page_size=pagination.page_size, results=items
    )


@router.get(
    "/students/select-options/",
    response=list[SelectOption],
    summary="Options etudiants pour dropdown",
)
def list_students_select(
    request,
    academic_year_id: int | None = None,
    search: str | None = None,
):
    qs = Student.objects.order_by("last_name", "first_name")
    if academic_year_id:
        qs = qs.filter(enrollments__academic_year_id=academic_year_id).distinct()
    if search:
        qs = qs.filter(
            Q(last_name__icontains=search)
            | Q(first_name__icontains=search)
            | Q(ine__icontains=search)
        )
    return [
        SelectOption(id=s.id, label=f"{s.last_name} {s.first_name} ({s.ine})")
        for s in qs[:500]
    ]


@router.get(
    "/students/stats/",
    response=StudentStatsOut,
    summary="Statistiques etudiants par annee",
)
def get_student_stats(request, academic_year_id: int):
    base = AnnualEnrollment.objects.filter(academic_year_id=academic_year_id)
    total = base.count()

    by_level = [
        LevelStatOut(
            level_id=item["level__id"],
            level_code=item["level__code"],
            level_name=item["level__name"],
            count=item["count"],
        )
        for item in base.values("level__id", "level__code", "level__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    ]

    by_department = [
        DepartmentStatOut(
            department_id=item["department__id"],
            department_code=item["department__code"],
            department_name=item["department__name"],
            count=item["count"],
        )
        for item in base.values(
            "department__id", "department__code", "department__name"
        )
        .annotate(count=Count("id"))
        .order_by("-count")
    ]

    by_parcours = [
        ParcoursStatOut(
            parcours_id=item["parcours__id"],
            parcours_code=item["parcours__code"],
            parcours_label=item["parcours__label"],
            count=item["count"],
        )
        for item in base.values("parcours__id", "parcours__code", "parcours__label")
        .annotate(count=Count("id"))
        .order_by("-count")
    ]

    cross = [
        CrossStatOut(
            level_id=item["level__id"],
            level_code=item["level__code"],
            level_name=item["level__name"],
            department_id=item["department__id"],
            department_code=item["department__code"],
            department_name=item["department__name"],
            parcours_id=item["parcours__id"],
            parcours_code=item["parcours__code"],
            parcours_label=item["parcours__label"],
            count=item["count"],
        )
        for item in base.values(
            "level__id",
            "level__code",
            "level__name",
            "department__id",
            "department__code",
            "department__name",
            "parcours__id",
            "parcours__code",
            "parcours__label",
        )
        .annotate(count=Count("id"))
        .order_by("level__code", "department__code", "parcours__code")
    ]

    return StudentStatsOut(
        total=total,
        by_level=by_level,
        by_department=by_department,
        by_parcours=by_parcours,
        cross=cross,
    )


@router.get("/students/template/", summary="Telecharger le template Excel etudiants")
def download_student_template(request):
    from app.reference.models import Parcours as ParcourModel

    dept_codes = list(
        Department.objects.filter(code__gt="")
        .values_list("code", flat=True)
        .order_by("code")
    )
    level_codes = list(
        Level.objects.filter(code__gt="")
        .values_list("code", flat=True)
        .order_by("code")
    )
    parcours_codes = list(
        ParcourModel.objects.values_list("code", flat=True).order_by("code").distinct()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etudiants"

    headers = [
        "INE",
        "Nom",
        "Prénom",
        "Email",
        "Genre",
        "Département",
        "Niveau",
        "Parcours",
        "GPA",
    ]
    header_fill = PatternFill(
        start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30
    for col, width in zip(
        "ABCDEFGHI", [15, 20, 20, 35, 10, 15, 12, 20, 10], strict=False
    ):
        ws.column_dimensions[col].width = width

    # Genre dropdown (column E) — inline list (short enough)
    dv_genre = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"M,F"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_genre.sqref = "E2:E10000"
    ws.add_data_validation(dv_genre)

    # Département dropdown (column F)
    if dept_codes:
        dept_ws = wb.create_sheet("_Departements")
        for i, code in enumerate(dept_codes, 1):
            dept_ws.cell(row=i, column=1, value=code)
        dept_ws.sheet_state = "hidden"
        dv_dept = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Departements!$A$1:$A${len(dept_codes)}",
            allow_blank=True,
        )
        dv_dept.sqref = "F2:F10000"
        ws.add_data_validation(dv_dept)

    # Niveau dropdown (column G)
    if level_codes:
        level_ws = wb.create_sheet("_Niveaux")
        for i, code in enumerate(level_codes, 1):
            level_ws.cell(row=i, column=1, value=code)
        level_ws.sheet_state = "hidden"
        dv_level = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Niveaux!$A$1:$A${len(level_codes)}",
            allow_blank=True,
        )
        dv_level.sqref = "G2:G10000"
        ws.add_data_validation(dv_level)

    # Parcours dropdown (column H)
    if parcours_codes:
        parcours_ws = wb.create_sheet("_Parcours")
        for i, code in enumerate(parcours_codes, 1):
            parcours_ws.cell(row=i, column=1, value=code)
        parcours_ws.sheet_state = "hidden"
        dv_parcours = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"_Parcours!$A$1:$A${len(parcours_codes)}",
            allow_blank=True,
        )
        dv_parcours.sqref = "H2:H10000"
        ws.add_data_validation(dv_parcours)

    # Example rows
    ws.append(
        [
            "123456789AB",
            "MARTIN",
            "Jean",
            "jean.martin@etud.n7.fr",
            "M",
            dept_codes[0] if dept_codes else "INFO",
            level_codes[0] if level_codes else "3A",
            parcours_codes[0] if parcours_codes else "SESG",
            15.5,
        ]
    )
    ws.append(
        [
            "987654321CD",
            "DUPONT",
            "Marie",
            "marie.dupont@etud.n7.fr",
            "F",
            dept_codes[1]
            if len(dept_codes) > 1
            else (dept_codes[0] if dept_codes else "TC"),
            level_codes[1]
            if len(level_codes) > 1
            else (level_codes[0] if level_codes else "4A"),
            parcours_codes[1]
            if len(parcours_codes) > 1
            else (parcours_codes[0] if parcours_codes else "IPA"),
            16.0,
        ]
    )

    # Instructions sheet
    info_ws = wb.create_sheet("Instructions")
    info_ws.column_dimensions["A"].width = 80
    instructions = [
        "INSTRUCTIONS IMPORT ETUDIANTS",
        "",
        "- Remplissez les donnees a partir de la ligne 2 (la ligne 1 est l'en-tete).",
        "- INE : identifiant national etudiant, 11 caracteres.",
        "- Genre : choisissez M (Homme) ou F (Femme) dans la liste deroulante.",
        "- Departement : choisissez dans la liste deroulante (codes de la base de donnees).",
        "- Niveau : choisissez dans la liste deroulante.",
        "- Parcours : choisissez dans la liste deroulante, ou laissez vide (tronc commun).",
        "- GPA : note moyenne sur 20, format decimal (ex : 15.5). Facultatif.",
        "",
        "Les lignes sans INE sont ignorees.",
        "Les departements ou niveaux introuvables en base sont signales dans le rapport d'import.",
    ]
    for i, line in enumerate(instructions, 1):
        cell = info_ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="template_etudiants.xlsx"'
    return response


@router.get(
    "/students/by-year/{year_id}/",
    response=PagedResponse[StudentEnrollmentOut],
    summary="Etudiants inscrits pour une annee avec details d'inscription",
)
def list_students_by_year(
    request,
    year_id: int,
    department_id: int | None = None,
    level_id: int | None = None,
    parcours_id: int | None = None,
    search: str | None = None,
    pagination: PaginationQuery = Query(),
):
    qs = (
        AnnualEnrollment.objects.filter(academic_year_id=year_id)
        .select_related(
            "student", "student__nationality", "department", "level", "parcours"
        )
        .order_by("student__last_name", "student__first_name")
    )
    if department_id:
        qs = qs.filter(department_id=department_id)
    if level_id:
        qs = qs.filter(level_id=level_id)
    if parcours_id:
        qs = qs.filter(parcours_id=parcours_id)
    if search:
        qs = qs.filter(
            Q(student__last_name__icontains=search)
            | Q(student__first_name__icontains=search)
            | Q(student__ine__icontains=search)
        )
    count, enrollments = paginate(qs, pagination.page, pagination.page_size)
    items = [
        StudentEnrollmentOut(
            student_id=e.student.id,
            ine=e.student.ine,
            first_name=e.student.first_name,
            last_name=e.student.last_name,
            email=e.student.email,
            gender=e.student.gender,
            nationality_iso2=e.student.nationality.iso2
            if e.student.nationality_id
            else None,
            nationality_name_fr=e.student.nationality.name_fr
            if e.student.nationality_id
            else None,
            department_id=e.department.id,
            department_code=e.department.code,
            department_name=e.department.name,
            level_id=e.level.id,
            level_code=e.level.code,
            level_name=e.level.name,
            parcours_id=e.parcours.id if e.parcours else None,
            parcours_code=e.parcours.code if e.parcours else None,
            parcours_label=e.parcours.label if e.parcours else None,
            gpa=e.gpa,
        )
        for e in enrollments
    ]
    return PagedResponse(
        count=count, page=pagination.page, page_size=pagination.page_size, results=items
    )


@router.get(
    "/students/import-errors/",
    response=list[StudentRawImportOut],
    summary="Erreurs d'import étudiants (Pégase ou Excel)",
)
def list_student_import_errors(request):
    raw_imports = RawImport.objects.filter(entity=RawImportEntity.STUDENT).order_by(
        "-created_at"
    )
    latest: dict = {}
    for ri in raw_imports:
        key = ri.external_id or f"raw-{ri.id}"
        if key not in latest:
            latest[key] = ri
    return [ri for ri in latest.values() if ri.status == RawImportStatus.FAILED]


@router.put(
    "/students/import-errors/{raw_import_id}/ignore/",
    response=StudentRawImportOut,
    summary="Marquer une erreur d'import étudiant comme traitée",
)
def ignore_student_import_error(request, raw_import_id: int):
    try:
        raw_import = RawImport.objects.get(
            pk=raw_import_id,
            entity=RawImportEntity.STUDENT,
            status=RawImportStatus.FAILED,
        )
    except RawImport.DoesNotExist as exc:
        raise HttpError(404, "Erreur d'import étudiant introuvable.") from exc

    raw_import.status = RawImportStatus.IGNORED
    raw_import.error_message = (
        f"{raw_import.error_message}\nTraité manuellement par l'administrateur."
    ).strip()
    raw_import.save(update_fields=["status", "error_message", "updated_at"])
    return raw_import


@router.put(
    "/students/import-errors/{raw_import_id}/retry/",
    response=StudentRawImportOut,
    summary="Relancer un import étudiant en corrigeant le département, niveau ou parcours",
)
def retry_student_import_error(
    request, raw_import_id: int, payload: StudentImportRetryIn
):
    try:
        raw_import = RawImport.objects.get(
            pk=raw_import_id,
            entity=RawImportEntity.STUDENT,
            status=RawImportStatus.FAILED,
        )
    except RawImport.DoesNotExist as exc:
        raise HttpError(404, "Erreur d'import étudiant introuvable.") from exc

    if raw_import.academic_year_id is None:
        raise HttpError(400, "Cet import n'est pas associé à une année universitaire.")

    corrected = dict(raw_import.payload)
    if payload.department_id is not None:
        dept = Department.objects.filter(pk=payload.department_id).first()
        if dept is None:
            raise HttpError(400, f"Département {payload.department_id} introuvable.")
        corrected["department_code"] = dept.code

    if payload.level_id is not None:
        level = Level.objects.filter(pk=payload.level_id).first()
        if level is None:
            raise HttpError(400, f"Niveau {payload.level_id} introuvable.")
        corrected["level_code"] = level.code

    if payload.parcours_id is not None:
        parcours = Parcours.objects.filter(pk=payload.parcours_id).first()
        if parcours is None:
            raise HttpError(400, f"Parcours {payload.parcours_id} introuvable.")
        corrected["parcours_code"] = parcours.code

    from app.academic.models import AcademicYear as AcademicYearModel

    academic_year = AcademicYearModel.objects.get(pk=raw_import.academic_year_id)

    try:
        ts = transform_student(corrected)
        validate_student(ts)
    except (ValueError, DomainValidationError) as exc:
        raise HttpError(400, str(exc)) from exc

    row = StudentRow(
        ine=ts.ine,
        first_name=ts.first_name,
        last_name=ts.last_name,
        email=ts.email,
        gender=ts.gender,
        department_code=ts.department_code,
        level_code=ts.level_code,
        parcours_code=ts.parcours_code,
        gpa=ts.gpa,
        nationality_iso2=ts.nationality_iso2,
    )

    report = import_students([row], academic_year)
    if report.errors or report.unresolved:
        reason = (report.errors or [str(report.unresolved[0])])[0]
        raw_import.payload = corrected
        raw_import.error_message = reason
        raw_import.save(update_fields=["payload", "error_message", "updated_at"])
        raise HttpError(400, reason)

    raw_import.payload = corrected
    raw_import.status = RawImportStatus.IMPORTED
    raw_import.error_message = ""
    from django.utils import timezone as tz

    raw_import.imported_at = tz.now()
    raw_import.save(
        update_fields=[
            "payload",
            "status",
            "error_message",
            "imported_at",
            "updated_at",
        ]
    )
    return raw_import


@router.put(
    "/students/wishes/import-errors/{raw_import_id}/retry/",
    response=StudentRawImportOut,
    summary="Relancer un import vœu en corrigeant l'étudiant ou l'accord",
)
def retry_wish_import_error(request, raw_import_id: int, payload: WishImportRetryIn):
    try:
        raw_import = RawImport.objects.get(
            pk=raw_import_id,
            entity=RawImportEntity.STUDENT,
            status=RawImportStatus.FAILED,
        )
    except RawImport.DoesNotExist as exc:
        raise HttpError(404, "Erreur d'import vœu introuvable.") from exc

    if raw_import.academic_year_id is None:
        raise HttpError(400, "Cet import n'est pas associé à une année universitaire.")

    corrected = dict(raw_import.payload)

    if payload.student_id is not None:
        student = Student.objects.filter(pk=payload.student_id).first()
        if student is None:
            raise HttpError(400, f"Étudiant {payload.student_id} introuvable.")
        corrected["ine"] = student.ine

    if payload.agreement_id is not None:
        agreement = Agreement.objects.filter(pk=payload.agreement_id).first()
        if agreement is None:
            raise HttpError(400, f"Accord {payload.agreement_id} introuvable.")
        corrected["offre_de_sejour"] = agreement.name

    from app.academic.models import AcademicYear as AcademicYearModel

    academic_year = AcademicYearModel.objects.get(pk=raw_import.academic_year_id)

    try:
        tw = transform_wish(corrected)
        validate_wish(tw)
    except (ValueError, DomainValidationError) as exc:
        raise HttpError(400, str(exc)) from exc

    row = WishRow(
        individu=tw.individu,
        offre_de_sejour=tw.offre_de_sejour,
        rank=tw.rank,
        ine=tw.ine,
    )

    report = import_wish_rows([row], academic_year)
    if report.errors or report.unresolved:
        reason = (report.errors or [str(report.unresolved[0])])[0]
        raw_import.payload = corrected
        raw_import.error_message = reason
        raw_import.save(update_fields=["payload", "error_message", "updated_at"])
        raise HttpError(400, reason)

    raw_import.payload = corrected
    raw_import.status = RawImportStatus.IMPORTED
    raw_import.error_message = ""
    from django.utils import timezone as tz

    raw_import.imported_at = tz.now()
    raw_import.save(
        update_fields=[
            "payload",
            "status",
            "error_message",
            "imported_at",
            "updated_at",
        ]
    )
    return raw_import


@router.get(
    "/students/wishes/template/{year_id}/",
    summary="Télécharger le template Excel vœux pour une année",
)
def download_wish_template(request, year_id: int):
    academic_year = get_academic_year(year_id)
    file_bytes = excel_wishes_adapter.generate_wish_template(academic_year)
    label_slug = academic_year.label.replace("/", "-")
    response = HttpResponse(
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="template_voeux_{label_slug}.xlsx"'
    )
    return response


@router.post(
    "/students/wishes/import-excel/{year_id}/",
    response=WishSyncReportOut,
    summary="Importer les vœux depuis un fichier Excel",
)
def import_wishes_from_excel(request, year_id: int, file: UploadedFile = File(...)):
    academic_year = get_academic_year(year_id)
    rows = excel_wishes_adapter.parse_wish_excel(file.read())
    db_report = DbImportReport.objects.create(
        source=ImportSource.EXCEL,
        academic_year=academic_year,
        triggered_by=getattr(request.user, "username", ""),
    )
    report = import_wish_rows(rows, academic_year, db_report=db_report)
    db_report.finalize()
    log_action(
        request,
        action="import_excel_wishes",
        detail=(
            f"Fichier {file.name} — Année {academic_year.label} — "
            f"{report.created} créés, {report.updated} mis à jour, "
            f"{len(report.unresolved)} non résolus"
        ),
    )
    return report


@router.get(
    "/students/export-excel/{year_id}/",
    summary="Exporter les étudiants en Excel",
)
def export_students_excel(
    request,
    year_id: int,
    level_id: int | None = None,
    dept_id: int | None = None,
    parcours_id: str | None = None,
):
    academic_year = get_academic_year(year_id)

    qs = (
        AnnualEnrollment.objects.filter(academic_year_id=year_id)
        .select_related(
            "student", "student__nationality", "department", "level", "parcours"
        )
        .order_by("student__last_name", "student__first_name")
    )
    if level_id:
        qs = qs.filter(level_id=level_id)
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if parcours_id == "none":
        qs = qs.filter(parcours__isnull=True)
    elif parcours_id:
        try:
            qs = qs.filter(parcours_id=int(parcours_id))
        except ValueError:
            pass

    # Construire le suffixe de nom de fichier
    label_slug = academic_year.label.replace("/", "-")
    dept_slug = ""
    level_slug = ""
    if dept_id:
        dept = Department.objects.filter(pk=dept_id).first()
        dept_slug = dept.code if dept else ""
    if level_id:
        lvl = Level.objects.filter(pk=level_id).first()
        level_slug = lvl.code if lvl else ""

    filename = build_filename("etudiants", label_slug, dept_slug, level_slug)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Étudiants"

    headers = [
        "INE",
        "Nom",
        "Prénom",
        "Email",
        "Genre",
        "Nationalité",
        "Département",
        "Niveau",
        "Parcours",
        "GPA",
    ]
    widths = [14, 22, 20, 32, 8, 20, 18, 14, 20, 8]
    write_header_row(ws, headers, widths)

    gender_labels = {"M": "Homme", "F": "Femme", "O": "Autre"}
    for e in qs:
        s = e.student
        ws.append(
            [
                s.ine,
                s.last_name,
                s.first_name,
                s.email,
                gender_labels.get(s.gender, ""),
                s.nationality.name_fr if s.nationality_id else "",
                f"{e.department.code}" if e.department else "",
                f"{e.level.code}" if e.level else "",
                f"{e.parcours.code}" if e.parcours else "",
                float(e.gpa) if e.gpa is not None else "",
            ]
        )

    return workbook_response(wb, filename)


@router.get(
    "/students/wishes/export-excel/{year_id}/",
    summary="Exporter les vœux en Excel",
)
def export_wishes_excel(
    request,
    year_id: int,
    dept_code: str | None = None,
):
    academic_year = get_academic_year(year_id)

    qs = (
        StudentWish.objects.filter(annual_enrollment__academic_year=academic_year)
        .select_related(
            "annual_enrollment__student",
            "annual_enrollment__department",
            "annual_enrollment__level",
            "agreement__partner_university__country",
        )
        .order_by(
            "annual_enrollment__student__last_name",
            "annual_enrollment__student__first_name",
            "rank",
        )
    )
    if dept_code:
        qs = qs.filter(annual_enrollment__department__code=dept_code)

    # Regrouper par étudiant
    rows: dict[int, dict] = {}
    for w in qs:
        enr = w.annual_enrollment
        sid = enr.student_id
        if sid not in rows:
            rows[sid] = {
                "ine": enr.student.ine,
                "nom": enr.student.last_name,
                "prenom": enr.student.first_name,
                "dept": enr.department.code if enr.department else "",
                "niveau": enr.level.code if enr.level else "",
                "wishes": [],
            }
        univ = (
            w.agreement.partner_university
            if w.agreement and w.agreement.partner_university_id
            else None
        )
        rows[sid]["wishes"].append(
            {
                "accord": w.agreement.name if w.agreement else "",
                "universite": univ.name if univ else "",
                "pays": univ.country.name_fr if univ and univ.country_id else "",
                "rank": w.rank,
            }
        )

    max_rank = max(
        (max((ww["rank"] for ww in r["wishes"]), default=0) for r in rows.values()),
        default=3,
    )
    max_rank = max(max_rank, 3)

    label_slug = academic_year.label.replace("/", "-")
    filename = build_filename("voeux", label_slug, dept_code or "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vœux"

    headers = ["INE", "Nom", "Prénom", "Département", "Niveau"]
    widths = [14, 22, 20, 14, 10]
    for r in range(1, max_rank + 1):
        headers.append(f"Vœu {r}")
        widths.append(70)
    write_header_row(ws, headers, widths)

    for r in sorted(rows.values(), key=lambda x: (x["nom"], x["prenom"])):
        wish_map = {w["rank"]: w for w in r["wishes"]}
        row_data = [r["ine"], r["nom"], r["prenom"], r["dept"], r["niveau"]]
        for rank in range(1, max_rank + 1):
            w = wish_map.get(rank)
            if w:
                parts = [p for p in [w["accord"], w["universite"], w["pays"]] if p]
                row_data.append(" — ".join(parts))
            else:
                row_data.append("")
        ws.append(row_data)

    return workbook_response(wb, filename)


@router.get(
    "/students/{student_id}/",
    response=StudentDetailOut,
    summary="Detail d'un etudiant",
)
def get_student(request, student_id: int):
    try:
        return (
            Student.objects.prefetch_related(
                Prefetch(
                    "enrollments",
                    queryset=AnnualEnrollment.objects.select_related(
                        "academic_year", "department", "level", "parcours"
                    ).order_by("-academic_year__start_date"),
                )
            )
            .select_related("nationality")
            .get(pk=student_id)
        )
    except Student.DoesNotExist as exc:
        raise HttpError(404, "Etudiant introuvable.") from exc


@router.post(
    "/students/sync-pegase/{year_id}/",
    response=ImportReportOut,
    summary="Synchroniser les etudiants depuis Pegase",
)
def sync_from_pegase(request, year_id: int):
    academic_year = get_academic_year(year_id)
    rows = pegase_adapter.fetch_enrollments(academic_year.label)
    db_report = DbImportReport.objects.create(
        source=ImportSource.PEGASE,
        academic_year=academic_year,
        triggered_by=getattr(request.user, "username", ""),
    )
    report = import_students(rows, academic_year, db_report=db_report, source_file="")
    db_report.finalize()
    log_action(
        request,
        action="sync_pegase_students",
        detail=f"Année {academic_year.label} — {report.created} créés, {report.updated} mis à jour, {len(report.unresolved)} non résolus",
    )
    return report


@router.post(
    "/students/import-excel/{year_id}/",
    response=ImportReportOut,
    summary="Importer les etudiants depuis un fichier Excel",
)
def import_from_excel(request, year_id: int, file: UploadedFile = File(...)):
    academic_year = get_academic_year(year_id)
    rows = excel_adapter.parse(file.read())
    db_report = DbImportReport.objects.create(
        source=ImportSource.EXCEL,
        academic_year=academic_year,
        triggered_by=getattr(request.user, "username", ""),
    )
    report = import_students(
        rows, academic_year, db_report=db_report, source_file=file.name
    )
    db_report.finalize()
    log_action(
        request,
        action="import_excel_students",
        detail=f"Fichier {file.name} — Année {academic_year.label} — {report.created} créés, {report.updated} mis à jour, {len(report.unresolved)} non résolus",
    )
    return report


# ── Vœux étudiants ─────────────────────────────────────────────────────────


@router.post(
    "/students/wishes/sync-moveon/{year_id}/",
    response=WishSyncReportOut,
    summary="Synchroniser les vœux étudiants depuis MoveON",
)
def sync_wishes_from_moveon(request, year_id: int):
    academic_year = get_academic_year(year_id)
    report = sync_moveon_wishes(
        academic_year=academic_year,
        triggered_by=getattr(request.user, "username", ""),
    )
    log_action(
        request,
        action="sync_moveon_wishes",
        detail=(
            f"Année {academic_year.label} — "
            f"{report.created} créés, {report.updated} mis à jour, "
            f"{len(report.unresolved)} non résolus"
        ),
    )
    return report


@router.get(
    "/students/wishes/by-year/{year_id}/",
    response=list[StudentWishesOut],
    summary="Vœux ordonnés par étudiant pour une année",
)
def list_wishes_by_year(request, year_id: int):
    academic_year = get_academic_year(year_id)

    wishes = (
        StudentWish.objects.filter(annual_enrollment__academic_year=academic_year)
        .select_related(
            "annual_enrollment__student",
            "annual_enrollment__department",
            "annual_enrollment__parcours",
            "agreement__partner_university",
        )
        .order_by(
            "annual_enrollment__student__last_name",
            "annual_enrollment__student__first_name",
            "rank",
        )
    )

    grouped: dict[int, StudentWishesOut] = {}
    for w in wishes:
        enrollment = w.annual_enrollment
        student = enrollment.student
        sid = student.id
        if sid not in grouped:
            grouped[sid] = StudentWishesOut(
                student_id=sid,
                ine=student.ine,
                first_name=student.first_name,
                last_name=student.last_name,
                department_code=enrollment.department.code,
                parcours_code=enrollment.parcours.code
                if enrollment.parcours_id
                else None,
                gpa=enrollment.gpa,
                wishes=[],
            )
        grouped[sid].wishes.append(
            AgreementWishOut(
                rank=w.rank,
                agreement_id=w.agreement_id,
                moveon_id=w.agreement.moveon_id,
                agreement_name=w.agreement.name,
                university_name=w.agreement.partner_university.name,
                direction=w.agreement.direction,
            )
        )

    return list(grouped.values())


@router.get(
    "/students/{student_id}/wishes/{year_id}/",
    response=StudentWishesOut,
    summary="Vœux d'un étudiant pour une année",
)
def get_student_wishes(request, student_id: int, year_id: int):
    try:
        student = Student.objects.get(pk=student_id)
    except Student.DoesNotExist as exc:
        raise HttpError(404, "Étudiant introuvable.") from exc

    academic_year = get_academic_year(year_id)

    enrollment = (
        AnnualEnrollment.objects.filter(student=student, academic_year=academic_year)
        .select_related("department", "parcours")
        .first()
    )

    wishes = (
        (
            StudentWish.objects.filter(annual_enrollment=enrollment)
            .select_related("agreement", "agreement__partner_university")
            .order_by("rank")
        )
        if enrollment
        else StudentWish.objects.none()
    )

    return StudentWishesOut(
        student_id=student.id,
        ine=student.ine,
        first_name=student.first_name,
        last_name=student.last_name,
        department_code=enrollment.department.code if enrollment else None,
        parcours_code=enrollment.parcours.code
        if enrollment and enrollment.parcours_id
        else None,
        gpa=enrollment.gpa if enrollment else None,
        wishes=[
            AgreementWishOut(
                rank=w.rank,
                agreement_id=w.agreement_id,
                moveon_id=w.agreement.moveon_id,
                agreement_name=w.agreement.name,
                university_name=w.agreement.partner_university.name,
                direction=w.agreement.direction,
            )
            for w in wishes
        ],
    )
