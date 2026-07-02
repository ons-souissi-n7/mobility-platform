"""
Adaptateur Excel pour les vœux de mobilité.

Format du template (1 ligne par étudiant) :
  Étudiant (INE – Nom Prénom) | Vœu 1 | Vœu 2 | Vœu 3

Colonne A : dropdown "INE – NOM Prénom" (sélection depuis la liste des étudiants inscrits).
Colonnes B-D : dropdown des accords actifs pour l'année.
Le parser extrait l'INE depuis la valeur "01234567890 – NOM Prénom" et génère
un WishRow par case Vœu remplie (rank = numéro de la colonne).
"""

import logging
import re
import unicodedata
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.academic.models import AcademicYear
from app.mobility.models import Agreement
from app.shared.excel_utils import format_agreement_label

from ...models import MAX_WISHES_PER_STUDENT
from ..sync_moveon import WishRow

logger = logging.getLogger(__name__)

MAX_WISHES = MAX_WISHES_PER_STUDENT  # nombre de colonnes Vœu dans le template


# ── helpers ───────────────────────────────────────────────────────────────────


def _normalize_header(h) -> str:
    if not h:
        return ""
    s = str(h).strip().lower()
    # Strip diacritics (é→e, à→a, etc.) via NFD decomposition
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    return (
        s.replace("œ", "oe")
        .replace(" ", "")
        .replace("_", "")
        .replace("(", "")
        .replace(")", "")
        .replace("'", "")
        .replace("'", "")
    )


def _wish_col_label(rank: int) -> str:
    return f"Vœu {rank}"


# ── template ──────────────────────────────────────────────────────────────────


def generate_wish_template(
    academic_year: AcademicYear, max_wishes: int = MAX_WISHES
) -> bytes:
    from app.students.models import AnnualEnrollment

    # Étudiants inscrits pour cette année
    enrollments = (
        AnnualEnrollment.objects.filter(academic_year=academic_year)
        .select_related("student", "department", "level")
        .order_by("student__last_name", "student__first_name")
    )
    student_rows = [
        (e.student.ine, e.student.last_name, e.student.first_name) for e in enrollments
    ]

    # Accords actifs pour cette année (fallback = tous)
    def _to_label(a: Agreement) -> str:
        univ = a.partner_university if a.partner_university_id else None
        return format_agreement_label(
            a.name,
            univ.name if univ else "",
            univ.country.name_fr if univ and univ.country_id else "",
        )

    _agreement_qs = (
        Agreement.objects.filter(
            year_instances__academic_year=academic_year,
            year_instances__is_active=True,
        )
        .distinct()
        .select_related("partner_university__country")
        .order_by("name")
    )
    agreements = [_to_label(a) for a in _agreement_qs]
    if not agreements:
        agreements = [
            _to_label(a)
            for a in Agreement.objects.select_related(
                "partner_university__country"
            ).order_by("name")
        ]

    wb = openpyxl.Workbook()

    # ── Feuille cachée : liste des étudiants (format "INE – NOM Prénom") ──
    ref_stu_ws = wb.create_sheet("_Etudiants")
    for ine, nom, prenom in student_rows:
        ref_stu_ws.append([f"{ine} – {nom} {prenom}"])
    ref_stu_ws.sheet_state = "hidden"

    # ── Feuille cachée : liste des accords ──
    ref_acc_ws = wb.create_sheet("_Accords")
    for i, name in enumerate(agreements, 1):
        ref_acc_ws.cell(row=i, column=1, value=name)
    ref_acc_ws.sheet_state = "hidden"

    # ── Feuille principale ──
    ws = wb.active
    ws.title = "Voeux"

    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    green_fill = PatternFill(
        start_color="166534", end_color="166534", fill_type="solid"
    )
    white_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    # En-têtes fixes
    fixed_headers = ["Étudiant (INE – Nom Prénom)"]
    fixed_widths = [45]

    # En-têtes vœux
    wish_headers = [_wish_col_label(r) for r in range(1, max_wishes + 1)]
    wish_widths = [45] * max_wishes

    all_headers = fixed_headers + wish_headers
    all_widths = fixed_widths + wish_widths

    for col, (header, width) in enumerate(
        zip(all_headers, all_widths, strict=False), 1
    ):
        is_wish = col > len(fixed_headers)
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = green_fill if is_wish else blue_fill
        cell.font = white_font
        cell.alignment = center
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    # Validation colonne A → liste "INE – NOM Prénom"
    if student_rows:
        n_stu = len(student_rows)
        dv_ine = DataValidation(
            type="list",
            formula1=f"_Etudiants!$A$1:$A${n_stu}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_ine.sqref = "A2:A10000"
        ws.add_data_validation(dv_ine)

    # Validation Vœu X (colonnes B…) → liste des accords
    if agreements:
        n_acc = len(agreements)
        for rank in range(1, max_wishes + 1):
            col_letter = openpyxl.utils.get_column_letter(len(fixed_headers) + rank)
            dv = DataValidation(
                type="list",
                formula1=f"_Accords!$A$1:$A${n_acc}",
                allow_blank=True,
                showDropDown=False,
            )
            dv.sqref = f"{col_letter}2:{col_letter}10000"
            ws.add_data_validation(dv)

    # Lignes d'exemple (2 premières) pré-remplies
    if student_rows and agreements:
        for idx in range(min(2, len(student_rows))):
            ine, nom, prenom = student_rows[idx]
            label = f"{ine} – {nom} {prenom}"
            row_data = [label] + [agreements[0]] + [""] * (max_wishes - 1)
            ws.append(row_data)

    # ── Feuille instructions ──
    info_ws = wb.create_sheet("Instructions")
    info_ws.column_dimensions["A"].width = 90
    last_col = openpyxl.utils.get_column_letter(len(fixed_headers) + max_wishes)
    lines = [
        f"IMPORT VŒUX DE MOBILITÉ — {academic_year.label}",
        "",
        "FORMAT : une ligne par étudiant, un accord par colonne Vœu.",
        "",
        "COLONNES :",
        "  A - Étudiant      : sélectionnez dans la liste déroulante (INE – Nom Prénom)",
        f"  B à {last_col} - Vœu 1 à {max_wishes} : sélectionnez un accord dans la liste déroulante",
        "",
        "RÈGLES :",
        "  - L'ordre des colonnes détermine la priorité : Vœu 1 = premier choix.",
        "  - Laissez une colonne Vœu vide si l'étudiant a moins de choix.",
        f"  - Maximum {max_wishes} vœux par étudiant.",
        "  - Les lignes sans étudiant sont ignorées.",
        f"  - L'étudiant doit être inscrit pour {academic_year.label} (sync Pégase d'abord).",
        "",
        "En cas d'erreur à l'import, un rapport détaillé est retourné.",
    ]
    for i, line in enumerate(lines, 1):
        cell = info_ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)
        elif (
            line.startswith("COLONNES")
            or line.startswith("RÈGLES")
            or line.startswith("FORMAT")
        ):
            cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ── parser ────────────────────────────────────────────────────────────────────


def parse_wish_excel(file_bytes: bytes) -> list[WishRow]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        return []

    # Détecter la colonne étudiant et les colonnes Vœu X
    student_col: int | None = None
    wish_cols: list[tuple[int, int]] = []  # (col_index, rank)

    for col_idx, h in enumerate(raw_headers):
        key = _normalize_header(h)
        # Colonne étudiant : "INE", "Étudiant", "etudiant(ine-nomprenom)", etc.
        if key in ("ine", "etudiant", "etudiante") or key.startswith("etudiant"):
            student_col = col_idx
        else:
            # Détecter "Vœu 1", "Voeu 2", "vœu3", "wish1", "choix1", etc.
            m = (
                re.match(r"v(?:o|oe|œ)eu(\d+)$", key)
                or re.match(r"wish(\d+)$", key)
                or re.match(r"choix(\d+)$", key)
            )
            if m:
                wish_cols.append((col_idx, int(m.group(1))))

    if student_col is None:
        logger.error("Colonne étudiant introuvable (attendu : 'Étudiant' ou 'INE')")
        return []

    if not wish_cols:
        logger.error("Aucune colonne Vœu trouvée (attendu : 'Vœu 1', 'Vœu 2', …)")
        return []

    wish_cols.sort(key=lambda x: x[1])

    result: list[WishRow] = []
    for raw_row in rows_iter:

        def get(idx: int | None, _row: list = raw_row) -> object:
            if idx is None or idx >= len(_row):
                return None
            return _row[idx]

        raw_student = str(get(student_col) or "").strip()
        if not raw_student:
            continue

        # Extraire l'INE depuis "01234567890 – NOM Prénom" ou valeur brute
        ine, individu = _extract_ine_and_name(raw_student)

        for col_idx, rank in wish_cols:
            accord = str(get(col_idx) or "").strip()
            if not accord:
                continue
            # Extrait le nom de l'accord depuis "Accord — Université — Pays"
            accord_name = accord.split(" — ")[0].strip() if " — " in accord else accord
            result.append(
                WishRow(
                    individu=individu,
                    offre_de_sejour=accord_name,
                    rank=rank,
                    ine=ine,
                )
            )

    return result


def _extract_ine_and_name(value: str) -> tuple[str, str]:
    """
    Extrait (ine, individu) depuis une valeur qui peut être :
      - "01234567890 – DUBOIS Lea"  (format template)
      - "01234567890 - DUBOIS Lea"  (tiret simple)
      - "01234567890"               (INE brut)
    """
    # Séparateur tiret long (–) ou tiret simple ( - )
    for sep in (" – ", " - ", "–", "-"):
        if sep in value:
            parts = value.split(sep, 1)
            ine = parts[0].strip()
            individu = parts[1].strip() if len(parts) > 1 else ""
            return ine, individu
    # Pas de séparateur : toute la valeur est traitée comme INE
    return value.strip(), ""
